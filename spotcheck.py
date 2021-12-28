#!/usr/bin/python3
########################################################### IMPORT MODULES - START #################################################################
from tkinter import *
from tkinter import messagebox
import time
from time import sleep, gmtime, strftime
from picamera import PiCamera
import cv2
import numpy as np
from tkinter import filedialog
from PIL import ImageTk, Image
import serial
from functools import partial
import math
from fractions import Fraction
from threading import *
import os
from tkinter import ttk
import awesometkinter as atk
import tkinter.font as font
import openpyxl
import subprocess
import shutil
import RPi.GPIO as GPIO
from ftplib import FTP
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image as Img
from datetime import *
############################################################ IMPORT MODULES - END ##################################################################

########################################################## GLOBAL VARIABLE - START #################################################################
covid19clicked = 0
viewresultclicked = 0
sorted_contours1 = list(range(48))
temp_label = 0
name = "/"
entry_num = 0
wait = 0
pos_result = list(range(48))
t2_tmp= list(range(48))
path0 = "/"
path1 = "/"
path2 = "/"
path3 = "/"
path4 = "/"
path5 = "/"
filename = ""
importfilename = ""
excel_file = ""
id_list = list(range(48))
covid19_createclicked = 0
samples = 0
covid19dir_old = ""
div = list(range(48))
start_point = (0,0)
end_point = (0,0)
t1_run = 0
t2_run = 0
t3_run = 0
t1_set = '65'
t2_set = '77'
t3_set = '80'
rsfile='/'
idfile='/'
test_list = list(range(48))
warning_value = 0
password = '123456789'
thr1_set = 1
thr2_set = 1
thr3l_set = 1
thr3h_set = 1

fr = open("/home/pi/Spotcheck/check.txt","r")
code = (fr.readline()).strip()
fr1 = open("/home/pi/Spotcheck/coordinates1.txt","r")
x1 = int(fr1.readline())
y1 = int(fr1.readline())
x2 = int(fr1.readline())
y2 = int(fr1.readline())
fr2 = open("/home/pi/Spotcheck/.server.txt","r")
server_on = int(fr2.readline())
ftp_ip = fr2.readline().strip('\n')
ftp_user = fr2.readline().strip('\n')
ftp_password = fr2.readline().strip('\n')
ftp_folder = fr2.readline().strip('\n')

hs = list(range(48))
workbook = openpyxl.load_workbook('/home/pi/Spotcheck/coefficient.xlsx')
sheet = workbook.active
for i in range(0,48):
    if(i<6):
        pos = str(chr(65+i+1)) + "2"
    if(i>=6 and i<12):
        pos = str(chr(65+i-5)) + "3"
    if(i>=12 and i<18):
        pos = str(chr(65+i-11)) + "4"
    if(i>=18 and i<24):
        pos = str(chr(65+i-17)) + "5"
    if(i>=24 and i<30):
        pos = str(chr(65+i-23)) + "6"
    if(i>=30 and i<36):
        pos = str(chr(65+i-29)) + "7"
    if(i>=36 and i<42):
        pos = str(chr(65+i-35)) + "8"
    if(i>=42):
        pos = str(chr(65+i-41)) + "9"
    hs[i] = float(sheet[pos].value)

fr3 = open("/var/tmp/.admin.txt","r")
start_trial = int(fr3.readline())
print("start_trial: ", start_trial)

fr4 = open("/home/pi/Spotcheck/mmvalue.txt","r")
v01 = float(fr4.readline())
v02= float(fr4.readline())

if not os.path.exists('/home/pi/Spotcheck Ket Qua'):
    f = os.path.join("/home/pi/", "Spotcheck Ket Qua")
    os.mkdir(f)
if not os.path.exists('/home/pi/Desktop/Spotcheck ID'):
    f = os.path.join("/home/pi/Desktop", "Spotcheck ID")
    os.mkdir(f)
if not os.path.exists('/home/pi/Desktop/Spotcheck ID/Spotcheck ID - Old'):
    f = os.path.join("/home/pi/Desktop/Spotcheck ID/", "Spotcheck ID - Old")
    os.mkdir(f)
if not os.path.exists('/home/pi/Desktop/Ket Qua Phan Tich'):
    f = os.path.join("/home/pi/Desktop/", "Ket Qua Phan Tich")
    os.mkdir(f)

########################################################### GLOBAL VARIABLE - END ##################################################################

################################################################# TRIAL _ START ####################################################################
def trial():
    old_day = int(fr3.readline())
    old_month = int(fr3.readline())
    old_year = int(fr3.readline())
    limit = int(fr3.readline())
    print("Ngày bắt đầu:", old_day, old_month, old_year)

    today = datetime.now()
    new_day = today.day
    new_month = today.month
    new_year = today.year
    print("Ngày hiện tại:", new_day, new_month, new_year)

    nam = new_year - old_year
    if(new_month < old_month):
        thang = new_month + 12 - old_month
        nam = nam - 1
    else:
        thang = new_month - old_month
    if(new_day < old_day):
        ngay = new_day + 30 - old_day
        thang = thang - 1
    else:
        ngay = new_day - old_day
    songay = ngay + thang*30 + nam*365
    print("Thời gian dùng thử còn lại:", limit-songay)
    if(songay >= limit):
        trial_labelframe = LabelFrame(root, bg='white', width=800, height=600)
        trial_labelframe.place(x=0,y=0)

        logo_img = Image.open('/home/pi/Spotcheck/logo.png')
        logo_width, logo_height = logo_img.size
        scale_percent = 50
        width = int(logo_width * scale_percent / 100)
        height = int(logo_height * scale_percent / 100)
        display_img = logo_img.resize((width,height))
        image_select = ImageTk.PhotoImage(display_img)
        logo_label = Label(trial_labelframe, bg='white',image=image_select)
        logo_label.image = image_select
        logo_label.place(x=5,y=5)

        def active_click(event = None):
            code = activecode_entry.get()
            if(code==""):
                msg = messagebox.showwarning(" ","Bạn chưa nhập mã kích hoạt !")
            else:
                if(code!=password):
                    msg = messagebox.showerror(" ","Mã kích hoạt không đúng !")
                if(code==password):
                    msg = messagebox.showinfo(" ","Kích hoạt thành công !")
                    f1=open("/var/tmp/.admin.txt",'w')
                    f1.writelines("0")
                    mainscreen()

        trial_label = Label(trial_labelframe, bg='white',fg="red", text="Thời gian dùng thử đã hết\nVui lòng nhập mã kích hoạt để tiếp tục sử dụng !", font=("Courier",18,"bold"))
        trial_label.place(x=62,y=85)
        contact_label = Label(trial_labelframe, bg='white', text="Liên hệ nhà cung cấp để nhận mã kích hoạt:", font=("Courier",12,"bold"))
        contact_label.place(x=73,y=435)
        mail_label = Label(trial_labelframe, bg='white', fg='blue',text="cskh@phusabiochem.com", font=("Courier",12,"bold"))
        mail_label.place(x=503,y=435)
        activecode_entry = Entry(trial_labelframe, width=27, bg='white', font=("Courier",14,"bold"))
        activecode_entry.place(x=246,y=215)
        activecode_entry.bind("<Return>", active_click)
        code_label = Label(trial_labelframe, bg='white', text="Mã kích hoạt:", font=("Courier",14,"bold"))
        code_label.place(x=244,y=189)

        key_img = Image.open('/home/pi/Spotcheck/key.png')
        logo_width, logo_height = key_img.size
        scale_percent = 5
        width = int(logo_width * scale_percent / 100)
        height = int(logo_height * scale_percent / 100)
        display_img = key_img.resize((width,height))
        image_select = ImageTk.PhotoImage(display_img)
        logo_label = Label(trial_labelframe, bg='white',image=image_select)
        logo_label.image = image_select
        logo_label.place(x=726,y=85)

        active_button = Button(trial_labelframe, bg="lavender", font=("Courier",11,'bold'), text="Xác nhận", height=3, width=10, borderwidth=0, command=active_click)
        active_button.place(x=340,y=260)
    else:
        mainscreen()
################################################################ TRIAL - END #######################################################################

########################################################## MAIN WINDOW INIT - START ################################################################
root = Tk()
root.title(" ")
root.geometry('1024x600')
root.configure(background = "white")
root.attributes('-fullscreen', True)
root.resizable(False,False)
def disable_event():
    pass
root.protocol("WM_DELETE_WINDOW", disable_event)
s = ttk.Style()
s.theme_use('clam')

if(code=='1111'):
    msgbox = messagebox.showerror(" ","Hệ thống lỗi, vui lòng liên hệ với nhà cung cấp !")
    if(msgbox=='ok'):
        root.destroy()

########################################################### MAIN WINDOW INIT - END #################################################################

########################################################### RESOURCE PATH - START ##################################################################
def resoure_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)
############################################################ RESOURCE PATH - END ###################################################################

############################################################ CAMERA INIT - START ###################################################################
def camera_capture(output):
    global stop_click
    stop_click = 0
    camera = PiCamera(framerate=Fraction(1,6), sensor_mode=3)
    camera.rotation = 180
    camera.iso = 200
    sleep(2)
    camera.shutter_speed = 6000000
    camera.exposure_mode = 'off'
    camera.capture(output)
    camera.close()
############################################################# CAMERA INIT - END ####################################################################

############################################################ SERIAL INIT - START ###################################################################
ser = serial.Serial(
    port = '/dev/serial0',
    baudrate = 115200,
    parity = serial.PARITY_NONE,
    stopbits = serial.STOPBITS_ONE,
    bytesize = serial.EIGHTBITS,
    timeout = 1
)
############################################################# SERIAL INIT - END ####################################################################

######################################################### SORTING CONTOURS - START #################################################################
def sorting_y(contour):
    rect_y = cv2.boundingRect(contour)
    return rect_y[1]
def sorting_x(contour):
    rect_x = cv2.boundingRect(contour)
    return rect_x[0]
def sorting_xy(contour):
    rect_xy = cv2.boundingRect(contour)
    return math.sqrt(math.pow(rect_xy[0],2) + math.pow(rect_xy[1],2))
########################################################## SORTING CONTOURS - END ##################################################################

########################################################## IMAGE ANALYSIS - START ##################################################################
def process_image(image_name, start_point=(x1,y1), end_point=(x2,y2)):
    global sorted_contours1
    image = cv2.imread(image_name)
    blur_img = cv2.GaussianBlur(image.copy(), (35,35), 0)
    gray_img = cv2.cvtColor(blur_img, cv2.COLOR_BGR2GRAY)
    thresh, binary_img = cv2.threshold(gray_img.copy(), 30, maxval=255, type=cv2.THRESH_BINARY)
    contours, hierarchy = cv2.findContours(binary_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    print("Number of contours: " + str(len(contours)))

    contours.sort(key=lambda data:sorting_xy(data))

    contour_img = np.zeros_like(gray_img)
    contour_img = cv2.rectangle(contour_img, start_point, end_point, (255,255,255), -1)
    rect_w = end_point[0] - start_point[0]
    rect_h = end_point[1] - start_point[1]
    cell_w = round(rect_w/6)
    cell_h = round(rect_h/8)
    for i in range(1,6):
        contour_img = cv2.line(contour_img, (start_point[0]+i*cell_w,start_point[1]), (start_point[0]+i*cell_w,end_point[1]),(0,0,0), 4)
    for i in range(1,8):
        contour_img = cv2.line(contour_img, (start_point[0],start_point[1]+i*cell_h), (end_point[0],start_point[1]+i*cell_h),(0,0,0), 4)

    thresh1 , binary1_img = cv2.threshold(contour_img, 250, maxval=255, type=cv2.THRESH_BINARY)
    contours1, hierarchy1 = cv2.findContours(binary1_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)

    contours1.sort(key=lambda data:sorting_y(data))
    contours1_h1 = contours1[0:6]
    contours1_h2 = contours1[6:12]
    contours1_h3 = contours1[12:18]
    contours1_h4 = contours1[18:24]
    contours1_h5 = contours1[24:30]
    contours1_h6 = contours1[30:36]
    contours1_h7 = contours1[36:42]
    contours1_h8 = contours1[42:48]
    contours1_h1.sort(key=lambda data:sorting_x(data))
    contours1_h2.sort(key=lambda data:sorting_x(data))
    contours1_h3.sort(key=lambda data:sorting_x(data))
    contours1_h4.sort(key=lambda data:sorting_x(data))
    contours1_h5.sort(key=lambda data:sorting_x(data))
    contours1_h6.sort(key=lambda data:sorting_x(data))
    contours1_h7.sort(key=lambda data:sorting_x(data))
    contours1_h8.sort(key=lambda data:sorting_x(data))

    sorted_contours1 = contours1_h1 + contours1_h2 + contours1_h3 + contours1_h4 + contours1_h5 + contours1_h6 + contours1_h7 + contours1_h8

    list_intensities = []
    sum_intensities = []
    result_list = list(range(48))
    area = list(range(48))

#Gray
    # tmp_list = list(range(48))
    # blur1_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
    # grayprocess_img = cv2.cvtColor(blur1_img, cv2.COLOR_BGR2GRAY)
    # #cv2.imwrite("mau.jpg",grayprocess_img)
    # for i in range(len(sorted_contours1)):
    #     cimg = np.zeros_like(gray_img)
    #     cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
    #     pts = np.where(cimg == 255)
    #     list_intensities.append(grayprocess_img[pts[0], pts[1]])
    #     list_intensities[i].sort()
    #     #print("list_intensities",str(i),":",list_intensities[i])
    #     #print("value", str(i), " : ", list_intensities[i][len(list_intensities[i])-1])
    #     sum_intensities.append(sum(list_intensities[i][len(list_intensities[i])-280:]))
    #     #sum_intensities.append(sum(list_intensities[i][len(list_intensities[i])-240:]))
    #     area[i]= cv2.contourArea(sorted_contours1[i])
    #     #result_list[i] = sum_intensities[i]
    #     tmp_list[i] = sum_intensities[i]/1000
    #     #result_list[i] = round(tmp_list[i])
    #     #result_list[i] = round(round(tmp_list[i],1)*1.5)
    #     result_list[i] = round(tmp_list[i],1)

#BGR
    blur1_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
    tmp_list = list(range(48))
    list_bgrvalue = []
    list_index = list(range(48))
    for i in range(len(sorted_contours1)):
        list_index[i] = []
        cimg = np.zeros_like(gray_img)
        cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
        pts = np.where(cimg == 255)
        list_bgrvalue.append(blur1_img[pts[0], pts[1]])
        for j in range(len(list_bgrvalue[i])):
             list_index[i].append(round((list_bgrvalue[i][j][0]+list_bgrvalue[i][j][1]+list_bgrvalue[i][j][2]/3)))
        list_index[i].sort()
        list_intensities.append(sum(list_index[i][len(list_index[i])-250:]))
        area[i]= cv2.contourArea(sorted_contours1[i])
        tmp_list[i] = list_intensities[i]/1000
        result_list[i] = round(tmp_list[i],1)

#HSV
#    tmp_list = list(range(48))
#     #blur1_img = cv2.fastNlMeansDenoisingColored(image.copy(),None,9,9,7,19)
#     #cv2.imwrite("mau1.jpg",blur1_img)
#     blur1_img = cv2.GaussianBlur(image.copy(), (3,3), 0)
#     cv2.imwrite("mau.jpg",blur_img)
#     hsv_img = cv2.cvtColor(blur1_img, cv2.COLOR_BGR2HSV)
#     list_hsvvalue = []
#     list_index = list(range(48))
#     for i in range(len(sorted_contours1)):
#         list_index[i] = []
#         cimg = np.zeros_like(gray_img)
#         cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
#         pts = np.where(cimg == 255)
#         list_hsvvalue.append(hsv_img[pts[0], pts[1]])
#         for j in range(len(list_hsvvalue[i])):
#             list_index[i].append(list_hsvvalue[i][j][2])
#         list_index[i].sort()
#         #print(len(list_index[i]))
#         list_intensities.append(sum(list_index[i][len(list_index[i])-250:]))
#         #area[i]= cv2.contourArea(sorted_contours1[i])
#         result_list[i] = list_intensities[i]
#         tmp_list[i] = list_intensities[i]/1000
#         result_list[i] = round(tmp_list[i])

#Nhân hệ số
    global hs
    for i in range(len(sorted_contours1)):
        result_list[i] = round(result_list[i]*hs[i],1)

#     for i in range(len(sorted_contours1)):
#         if(i==0):
#             result_list[i] = round(result_list[i]*1.03,1)
#         if(i==47):
#             result_list[i] = round(result_list[i]*0.92,1)
#         if(i==37 or i==38 or i==39 or i==40 or
#            i==43 or i==44 or i==45 or
#            i==10 or i==16 or i==22 or i==28 or i==34):
#             result_list[i] = round(result_list[i]*0.98,1)
#         if(i==36 or i==42 or i==46 or i==4 or
#            i==5 or i==11 or i==17 or i==23 or i==29 or i==35 or i==41):
#             result_list[i] = round(result_list[i]*0.96,1)

    for i in range(len(sorted_contours1)):
        if(result_list[i]>99):
            result_list[i]=99

    for i in range(len(sorted_contours1)):
        if ((i!=0) and ((i+1)%6==0)):
            print('%.1f'%(result_list[i]))
        else:
            print('%.1f'%(result_list[i]), end = ' | ')

    blurori_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
    global t1_run, t2_run, t3_run, thr1_set, thr2_set, thr3l_set, id_list
    for i in range(len(sorted_contours1)):
        if(id_list[i]=='N/A'):
            cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,0), thickness = -1)
        else:
            if(t1_run==0 and t2_run==0 and t3_run==0):
                if(result_list[i]<=thr1_set):
                    cv2.drawContours(blurori_img, sorted_contours1, i, (0,255,0), thickness = 2)
                else:
                    cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,255), thickness = 2)

            else:
                if(t1_run==1):
                    if(result_list[i] <= float(thr1_set)):
                        cv2.drawContours(blurori_img, sorted_contours1, i, (0,255,0), thickness = 2)
                    else:
                        cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,255), thickness = 2)
                if(t2_run==1):
                    if(result_list[i] <= float(thr2_set)):
                        cv2.drawContours(blurori_img, sorted_contours1, i, (0,255,0), thickness = 2)
                    else:
                        cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,255), thickness = 2)
                if(t3_run==1):
                    if(result_list[i] <= float(thr3l_set)):
                        cv2.drawContours(blurori_img, sorted_contours1, i, (0,255,0), thickness = 2)
                    else:
                        cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,255), thickness = 2)                            
                                  
    return (result_list, blurori_img)
########################################################### IMAGE ANALYSIS - END ###################################################################

############################################################ MAIN SCREEN - START ###################################################################
def mainscreen():
    buttonFont = font.Font(family='Helvetica', size=10, weight='bold')
    global mainscreen_labelframe
    mainscreen_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    mainscreen_labelframe.place(x=0,y=0)
    sidebar_labelframe = LabelFrame(mainscreen_labelframe, font=("Courier",15,'bold'), bg='dodger blue', width=170, height=478)
    sidebar_labelframe.place(x=0,y=0)

    def home_click():
        try:
            subprocess.Popen(['killall','florence'])
        except:
            pass
        root.attributes('-fullscreen', True)

        home_canvas['bg'] = 'white'
        covid19_canvas['bg'] = 'dodger blue'
        viewresult_canvas['bg'] = 'dodger blue'
        setid_canvas['bg'] = 'dodger blue'
        config_canvas['bg'] = 'dodger blue'
        power_canvas['bg'] = 'dodger blue'

        global covid19clicked
        covid19clicked = 0

        global covid19_createclicked
        covid19_createclicked = 0

        homemc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        homemc_labelframe.place(x=172,y=0)

        logo_img = Image.open('/home/pi/Spotcheck/logo.png')
        logo_width, logo_height = logo_img.size
        scale_percent = 50
        width = int(logo_width * scale_percent / 100)
        height = int(logo_height * scale_percent / 100)
        display_img = logo_img.resize((width,height))
        image_select = ImageTk.PhotoImage(display_img)
        logo_label = Label(mainscreen_labelframe, bg='white',image=image_select)
        logo_label.image = image_select
        logo_label.place(x=558,y=10)

        sc48_label = Label(mainscreen_labelframe, font=("Courier",45,'bold'), bg='white', fg='dodger blue', text="SPOTCHECK-SC48")
        sc48_label.place(x=229, y=160)

        transform_label = Label(mainscreen_labelframe, font=("Courier",16,'bold'), bg='white', fg='red', text="TRANSFORM YOUR PCR INTO REAL-TIME MODE")
        transform_label.place(x=233, y=227)

        covidapp_label = Label(mainscreen_labelframe, font=("Courier",25,'bold'), bg='white', fg='grey30', text="COVID-19 APPLICATION")
        covidapp_label.place(x=280, y=270)

        if(warning_value==1):
            warning_label = Label(mainscreen_labelframe, bg='red',text='Hệ thống đang tản nhiệt, không đặt mẫu vào lúc này !', font=("Courier", 13, 'bold'))
            warning_label.place(x=220,y=450)
        else:
            warning_label = Label(mainscreen_labelframe, bg='white', fg='white', text='Hệ thống đang tản nhiệt, không đặt mẫu vào lúc này !', font=("Courier", 13, 'bold'))
            warning_label.place(x=220,y=450)

    def covid19_click():
        home_canvas['bg'] = 'dodger blue'
        covid19_canvas['bg'] = 'white'
        viewresult_canvas['bg'] = 'dodger blue'
        setid_canvas['bg'] = 'dodger blue'
        config_canvas['bg'] = 'dodger blue'
        power_canvas['bg'] = 'dodger blue'

        global covid19clicked
        covid19clicked = 1

        global spotcheck_createclicked, tb_createclicked, shrimp_createclicked
        spotcheck_createclicked = 0

        covid19mc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        covid19mc_labelframe.place(x=172,y=0)

        enterframe_labelframe = LabelFrame(covid19mc_labelframe, bg='white', width=609, height=175)
        enterframe_labelframe.place(x=5,y=5)

        folder1name_label = Label(enterframe_labelframe, bg='dodger blue',width=75, height=1)
        folder1name_label.place(x=0,y=1)

        foldername_label = Label(enterframe_labelframe, bg='white',text='Tệp mẫu xét nghiệm', fg='black', font=("Courier",12,'bold'))
        foldername_label.place(x=75,y=37)

        # import_label = Label(enterframe_labelframe, bg='white',text='ID file', fg='black', font=("Courier",11,'bold'))
        # import_label.place(x=68,y=80)

        file_label = Label(enterframe_labelframe, bg='white', fg='grey25', font=("Courier",13,'bold'))
        file_label.place(x=213,y=87)

        global covid19_createclicked
        global covid19dir_old
        global importfilename
        directory = strftime("COVID19 %y-%m-%d %H.%M.%S")
        if(covid19_createclicked == 0):
            directory_label = Label(enterframe_labelframe, font=("Courier",10,'bold'), bg='dodger blue', text=directory)
            directory_label.place(x=184,y=1)
            covid19dir_old = directory
        else:
            directory_label = Label(enterframe_labelframe, font=("Courier",10,'bold'), bg='dodger blue', text=covid19dir_old)
            directory_label.place(x=184,y=1)
            file_label['text'] = importfilename

        def thread():
            th1 = Thread(target = create_click)
            th1.start()
        def create_click(event=None):
            global covid19_createclicked
            covid19_createclicked = 1
            global foldername
            #foldername = str(file_label['text'])
            name = strftime(importfilename)
            global path0
            global covid19dir_old
            print("covid19dir_old:",covid19dir_old)
            path0 = os.path.join("/home/pi/Spotcheck Ket Qua/", covid19dir_old +" "+ name +"/")

            if(file_label['text']==""):
                msgbox = messagebox.showwarning(" ","Bạn chưa tải tệp lên !" )
            else:
                if os.path.exists(path0):
                    msg = messagebox.askquestion("Thư mục đã tồn tại", "Bạn có muốn thay thế thư mục cũ ?")
                    if(msg=='yes'):
                        shutil.rmtree(path0)
                        os.mkdir(path0)
                        global path1
                        path1 = os.path.join(path0,"Ảnh chụp")
                        os.mkdir(path1)
                        global path2
                        path2 = os.path.join(path0,"Ảnh xử lý")
                        os.mkdir(path2)
                        global path3
                        path3 = os.path.join(path0,"Bảng kết quả")
                        os.mkdir(path3)
                        global path4
                        path4 = os.path.join(path0,"Ảnh nguyên mẫu")
                        os.mkdir(path4)
                        global path5
                        path5 = os.path.join(path0,"Chương trình nhiệt")
                        os.mkdir(path5)
                        mainscreen_labelframe.place_forget()
                        scanposition()
                else:
                    os.mkdir(path0)
                    path1 = os.path.join(path0,"Ảnh chụp")
                    os.mkdir(path1)
                    path2 = os.path.join(path0,"Ảnh xử lý")
                    os.mkdir(path2)
                    path3 = os.path.join(path0,"Bảng kết quả")
                    os.mkdir(path3)
                    path4 = os.path.join(path0,"Ảnh nguyên mẫu")
                    os.mkdir(path4)
                    path5 = os.path.join(path0,"Chương trình nhiệt")
                    os.mkdir(path5)

                    global thr1_set, thr2_set,thr3l_set, thr3h_set
                    fr3 = open("/home/pi/Spotcheck/ct.txt","r")
                    thr1_set = float(fr3.readline())
                    thr2_set = float(fr3.readline())
                    thr3l_set = float(fr3.readline())
                    #thr3h_set = float(fr3.readline())

                    mainscreen_labelframe.place_forget()
                    scanposition()

        def import_click():
            if(server_on==1):
                try:
                    ftp = FTP(ftp_ip, ftp_user, ftp_password)
                    ftp.cwd(ftp_folder + 'UnProcessed_Data')
                    ftpfiles = ftp.nlst()
                    for ftpfile in ftpfiles:
                        if(os.path.exists("/home/pi/Desktop/Spotcheck ID/" + ftpfile)):
                            pass
                        elif(os.path.exists("/home/pi/Desktop/Spotcheck ID/Spotcheck ID - Old/" + ftpfile)):
                            pass
                        else:
                            localfolder = os.path.join('/home/pi/Desktop/Spotcheck ID/', ftpfile)
                            file = open(localfolder,'wb')
                            ftp.retrbinary('RETR ' + ftpfile, file.write)
                            file.close()
                            print(ftpfile, "download done!")
                    ftp.quit()
                except Exception as e :
                    error = messagebox.showwarning("Có lỗi xảy ra khi đồng bộ server !",str(e))
                    if(error=='ok'):
                        pass
            file = filedialog.askopenfile(initialdir='/home/pi/Desktop/Spotcheck ID/', mode='r', filetypes=[('Excel file','*.xlsm *.xlsx *.xls')])
            global importfilename
            global filename
            filename = file.name
            global excel_file
            if file is not None:
                a=0
                for i in range(len(filename)):
                    if(filename[i]=='/'):
                        a=i+1
                importfilename = filename[a:(len(filename)-5)]
                if (os.path.exists("/home/pi/Desktop/Ket Qua Phan Tich/" + importfilename + ".xlsm")):
                    messagebox.showwarning("","Tệp vừa chọn đã được sử dụng !")
                    create_button['state']='disabled'
                    create_button['bg']='grey75'
                    file_label['text'] = ""
                    #import_click()
                else:
                    excel_file = filename[a:len(filename)]
                    if(len(importfilename)>=15):
                        file_label['text'] = importfilename[0:15] + '...'
                    else:
                        file_label['text'] = importfilename

                    workbook = openpyxl.load_workbook(filename)
                    sheet = workbook.active
                    # for i in range(0,48):
                    #     if(i<6):
                    #         pos = str(chr(65+i+1)) + "2"
                    #     if(i>=6 and i<12):
                    #         pos = str(chr(65+i-5)) + "3"
                    #     if(i>=12 and i<18):
                    #         pos = str(chr(65+i-11)) + "4"
                    #     if(i>=18 and i<24):
                    #         pos = str(chr(65+i-17)) + "5"
                    #     if(i>=24 and i<30):
                    #         pos = str(chr(65+i-23)) + "6"
                    #     if(i>=30 and i<36):
                    #         pos = str(chr(65+i-29)) + "7"
                    #     if(i>=36 and i<42):
                    #         pos = str(chr(65+i-35)) + "8"
                    #     if(i>=42):
                    #         pos = str(chr(65+i-41)) + "9"

                    tmp_list = list(range(48))
                    for i in range(0,48):
                        pos = "B" + str(i+12)
                        tmp_list[i] = sheet[pos].value
                        if(i==0):
                            id_list[0] = tmp_list[i]
                        if(i==1):
                            id_list[6] = tmp_list[i]
                        if(i==2):
                            id_list[12] = tmp_list[i]
                        if(i==3):
                            id_list[18] = tmp_list[i]
                        if(i==4):
                            id_list[24] = tmp_list[i]
                        if(i==5):
                            id_list[30] = tmp_list[i]
                        if(i==6):
                            id_list[36] = tmp_list[i]
                        if(i==7):
                            id_list[42] = tmp_list[i]
                        if(i==8):
                            id_list[1] = tmp_list[i]
                        if(i==9):
                            id_list[7] = tmp_list[i]
                        if(i==10):
                            id_list[13] = tmp_list[i]
                        if(i==11):
                            id_list[19] = tmp_list[i]
                        if(i==12):
                            id_list[25] = tmp_list[i]
                        if(i==13):
                            id_list[31] = tmp_list[i]
                        if(i==14):
                            id_list[37] = tmp_list[i]
                        if(i==15):
                            id_list[43] = tmp_list[i]
                        if(i==16):
                            id_list[2] = tmp_list[i]
                        if(i==17):
                            id_list[8] = tmp_list[i]
                        if(i==18):
                            id_list[14] = tmp_list[i]
                        if(i==19):
                            id_list[20] = tmp_list[i]
                        if(i==20):
                            id_list[26] = tmp_list[i]
                        if(i==21):
                            id_list[32] = tmp_list[i]
                        if(i==22):
                            id_list[38] = tmp_list[i]
                        if(i==23):
                            id_list[44] = tmp_list[i]
                        if(i==24):
                            id_list[3] = tmp_list[i]
                        if(i==25):
                            id_list[9] = tmp_list[i]
                        if(i==26):
                            id_list[15] = tmp_list[i]
                        if(i==27):
                            id_list[21] = tmp_list[i]
                        if(i==28):
                            id_list[27] = tmp_list[i]
                        if(i==29):
                            id_list[33] = tmp_list[i]
                        if(i==30):
                            id_list[39] = tmp_list[i]
                        if(i==31):
                            id_list[45] = tmp_list[i]
                        if(i==32):
                            id_list[4] = tmp_list[i]
                        if(i==33):
                            id_list[10] = tmp_list[i]
                        if(i==34):
                            id_list[16] = tmp_list[i]
                        if(i==35):
                            id_list[22] = tmp_list[i]
                        if(i==36):
                            id_list[28] = tmp_list[i]
                        if(i==37):
                            id_list[34] = tmp_list[i]
                        if(i==38):
                            id_list[40] = tmp_list[i]
                        if(i==39):
                            id_list[46] = tmp_list[i]
                        if(i==40):
                            id_list[5] = tmp_list[i]
                        if(i==41):
                            id_list[11] = tmp_list[i]
                        if(i==42):
                            id_list[17] = tmp_list[i]
                        if(i==43):
                            id_list[23] = tmp_list[i]
                        if(i==44):
                            id_list[29] = tmp_list[i]
                        if(i==45):
                            id_list[35] = tmp_list[i]
                        if(i==46):
                            id_list[41] = tmp_list[i]
                        if(i==47):
                            id_list[47] = tmp_list[i]

                    create_button['state']='normal'
                    create_button['bg']='lawn green'

        import_button = Button(enterframe_labelframe, font=("Courier",12,'bold'), bg="lavender", text="Tải lên", height=3, width=10, borderwidth=0, command=import_click)
        import_button.place(x=78,y=64)

        if(file_label['text']==""):
            create_button = Button(enterframe_labelframe, font=("Courier",12,'bold'), bg="grey75", text="Tiếp theo", height=3, width=10, borderwidth=0, command=thread, state='disabled')
        else:
            create_button = Button(enterframe_labelframe, font=("Courier",12,'bold'), bg="lawn green", text="Tiếp theo", height=3, width=10, borderwidth=0, command=thread)
        create_button.place(x=403,y=64)

        if(warning_value==1):
            warning_label = Label(mainscreen_labelframe, bg='red',text='Hệ thống đang tản nhiệt, không đặt mẫu vào lúc này !', font=("Courier", 13, 'bold'))
            warning_label.place(x=220,y=450)
        else:
            warning_label = Label(mainscreen_labelframe, bg='white', fg='white', text='Hệ thống đang tản nhiệt, không đặt mẫu vào lúc này !', font=("Courier", 13, 'bold'))
            warning_label.place(x=220,y=450)

    def setid_click():
        home_canvas['bg'] = 'dodger blue'
        covid19_canvas['bg'] = 'dodger blue'
        viewresult_canvas['bg'] = 'dodger blue'
        setid_canvas['bg'] = 'white'
        config_canvas['bg'] = 'dodger blue'
        power_canvas['bg'] = 'dodger blue'

        setidmc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        setidmc_labelframe.place(x=172,y=0)

        setid_labelframe = LabelFrame(setidmc_labelframe, bg='white', width=430, height=435)
        setid_labelframe.place(x=95,y=5)

        setid0_label = Label(setid_labelframe, bg='dodger blue', text=" VỊ TRÍ ĐẶT MẪU", font=("Courier", 12,'bold'),width=42)
        setid0_label.place(x=1,y=1)

        s48_img = Image.open('/home/pi/Spotcheck/48well.JPG')
        s48_width, s48_height = s48_img.size
        scale_percent = 56
        width = int(s48_width * scale_percent / 100)
        height = int(s48_height * scale_percent / 100)
        display_img = s48_img.resize((width,height))
        image_select = ImageTk.PhotoImage(display_img)
        setid_label = Label(setid_labelframe, bg='white',image=image_select)
        setid_label.image = image_select
        setid_label.place(x=99,y=40)

        def ok_click():
            global covid19_createclicked
            covid19_createclicked = 0
            setid()

        ok_button = Button(setid_labelframe, fg='black', font=('Courier','13','bold'), bg="lavender", text="Tiếp theo", height=2, width=10, borderwidth=0, command=ok_click)
        ok_button.place(x=150,y=355)

        if(warning_value==1):
            warning_label = Label(mainscreen_labelframe, bg='red',text='Hệ thống đang tản nhiệt, không đặt mẫu vào lúc này !', font=("Courier", 13, 'bold'))
            warning_label.place(x=220,y=450)
        else:
            warning_label = Label(mainscreen_labelframe, bg='white', fg='white', text='Hệ thống đang tản nhiệt, không đặt mẫu vào lúc này !', font=("Courier", 13, 'bold'))
            warning_label.place(x=220,y=450)

    def power_click():
        home_canvas['bg'] = 'dodger blue'
        covid19_canvas['bg'] = 'dodger blue'
        viewresult_canvas['bg'] = 'dodger blue'
        setid_canvas['bg'] = 'dodger blue'
        config_canvas['bg'] = 'dodger blue'
        power_canvas['bg'] = 'white'

        powermc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        powermc_labelframe.place(x=172,y=0)

        power_labelframe = LabelFrame(powermc_labelframe, bg='white', width=405, height=120)
        power_labelframe.place(x=106,y=200)
        def shutdown_click():
            os.system("sudo shutdown -h now")
        def restart_click():
            os.system("sudo shutdown -r now")
        def exit_click():
            root.destroy()
        exit_button = Button(power_labelframe, fg='white', activebackground="dodger blue", font=('Courier','10','bold'), bg="blue4", text="Đóng ứng dụng", height=5, width=12, borderwidth=0, command=exit_click)
        exit_button.place(x=9,y=12)
        shutdown_button = Button(power_labelframe, fg='white', activebackground="red", font=('Courier','10','bold'), bg="red3", text="Tắt nguồn", height=5, width=12, borderwidth=0, command=shutdown_click)
        shutdown_button.place(x=139,y=12)
        restart_button = Button(power_labelframe, fg='white', activebackground="lawn green", font=('Courier','10','bold'), bg="green", text="Khởi động lại", height=5, width=12, borderwidth=0, command=restart_click)
        restart_button.place(x=269,y=12)

        if(warning_value==1):
            warning_label = Label(mainscreen_labelframe, bg='red',text='Hệ thống đang tản nhiệt, không đặt mẫu vào lúc này !', font=("Courier", 13, 'bold'))
            warning_label.place(x=220,y=450)
        else:
            warning_label = Label(mainscreen_labelframe, bg='white', fg='white', text='Hệ thống đang tản nhiệt, không đặt mẫu vào lúc này !', font=("Courier", 13, 'bold'))
            warning_label.place(x=220,y=450)

    def viewresult_click():
        home_canvas['bg'] = 'dodger blue'
        covid19_canvas['bg'] = 'dodger blue'
        viewresult_canvas['bg'] = 'white'
        setid_canvas['bg'] = 'dodger blue'
        config_canvas['bg'] = 'dodger blue'
        power_canvas['bg'] = 'dodger blue'

        viewresultmc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        viewresultmc_labelframe.place(x=172,y=0)

        viewresult_labelframe = LabelFrame(viewresultmc_labelframe, bg='white', width=299, height=160)
        viewresult_labelframe.place(x=156,y=152)

        viewresult_label = Label(viewresult_labelframe, bg='dodger blue', text="Kết quả", font=('Courier',13,'bold'), width=29, height=1)
        viewresult_label.place(x=0,y=1)

        def open_click():
            global rsfile
            rsfile = filedialog.askopenfilename(initialdir='/home/pi/Spotcheck Ket Qua/',filetypes=[('jpg file','*.jpg')])
            if rsfile is not None:
                if(rsfile[len(rsfile)-3:]=='jpg'):
                    global covid19_createclicked
                    covid19_createclicked = 0
                    print(rsfile)
                    result()
                else:
                    pass

        open_button = Button(viewresultmc_labelframe, bg="lavender", text="Mở", borderwidth=0, font=('Courier',13,'bold'), height=3, width=12, command=open_click)
        open_button.place(x=233,y=205)

        if(warning_value==1):
            warning_label = Label(mainscreen_labelframe, bg='red',text='Hệ thống đang tản nhiệt, không đặt mẫu vào lúc này !', font=("Courier", 13, 'bold'))
            warning_label.place(x=220,y=450)
        else:
            warning_label = Label(mainscreen_labelframe, bg='white', fg='white', text='Hệ thống đang tản nhiệt, không đặt mẫu vào lúc này !', font=("Courier", 13, 'bold'))
            warning_label.place(x=220,y=450)

    def config_click():
        home_canvas['bg'] = 'dodger blue'
        covid19_canvas['bg'] = 'dodger blue'
        viewresult_canvas['bg'] = 'dodger blue'
        setid_canvas['bg'] = 'dodger blue'
        config_canvas['bg'] = 'white'
        power_canvas['bg'] = 'dodger blue'

        configmc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        configmc_labelframe.place(x=172,y=0)

        def ct_click():
            configmc1_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
            configmc1_labelframe.place(x=172,y=0)
            config1_lableframe = LabelFrame(configmc1_labelframe, bg='white', text="Loại kit ly trích", width=402, height=120)
            config1_lableframe.place(x=107,y=130)
            ct_label = Label(configmc1_labelframe, text='CHỌN KIT LY TRÍCH',font=('bold'), width=61, bg='dodger blue')
            ct_label.place(x=3,y=1)

            fr4 = open("/home/pi/Spotcheck/ct.txt","r")
            firstline = (fr4.readline()).strip()
            secondline = (fr4.readline()).strip()
            thirdline = (fr4.readline()).strip()
            fourthline = (fr4.readline()).strip()

            def kit1_click():
                kit1_button['bg'] = 'lawn green'
                kit2_button['bg'] = 'grey88'
                kit1_button['fg'] = 'black'
                kit2_button['fg'] = 'grey70'

            def kit2_click():
                kit1_button['bg'] = 'grey88'
                kit2_button['bg'] = 'lawn green'
                kit1_button['fg'] = 'grey70'
                kit2_button['fg'] = 'black'

            var = IntVar()
            radio1 = Radiobutton(configmc1_labelframe, bg='white', width=19, font=('Courier',15), borderwidth=0, text="Ct ≤ 30 (Điều trị)", variable=var, value=1)
            radio1.place(x=175,y=250)
            radio2 = Radiobutton(configmc1_labelframe, bg='white', width=19, font=('Courier',15), borderwidth=0, text="Ct > 30 (Tầm soát)", variable=var, value=2)
            radio2.place(x=175,y=277)

            if(thirdline=='7.5' or thirdline=='7.3'):
                kit1_button = Button(config1_lableframe, bg="lawn green", text="Kit ly trích Phù Sa", font=("Helvetica",12, 'bold'), borderwidth=0, height=4, width=17, command=kit1_click)
                kit1_button.place(x=8,y=2)
                kit2_button = Button(config1_lableframe, bg="grey88", fg='grey70', text="Kit ly trích khác", font=("Helvetica",12,'bold'), borderwidth=0, height=4, width=17, command=kit2_click)
                kit2_button.place(x=210,y=2)
                if(thirdline=='7.5'):
                    radio1.select()
                else:
                    radio2.select()

            else:
                kit1_button = Button(config1_lableframe, bg="grey88", fg='grey70', text="Kit ly trích Phù Sa", font=("Helvetica",12, 'bold'), borderwidth=0, height=4, width=17, command=kit1_click)
                kit1_button.place(x=8,y=2)
                kit2_button = Button(config1_lableframe, bg="lawn green", text="Kit ly trích khác", font=("Helvetica",12,'bold'), borderwidth=0, height=4, width=17, command=kit2_click)
                kit2_button.place(x=210,y=2)
                if(thirdline=='7.8'):
                    radio1.select()
                else:
                    radio2.select()

            def save_click():
                msg = messagebox.askquestion("Lưu ", "Bạn có muốn lưu lựa chọn ?")
                if(msg=='yes'):
                    radio_select = var.get()
                    if(radio_select==1 and kit1_button['bg']=='lawn green'):
                        tc= open("/home/pi/Spotcheck/ct.txt","w")
                        tc.truncate(0)
                        tc.writelines("7.0"+"\n")
                        tc.writelines("7.0"+"\n")
                        tc.writelines("7.5"+"\n")
                        tc.writelines("7.8"+"\n")
                    if(radio_select==2 and kit1_button['bg']=='lawn green'):
                        tc= open("/home/pi/Spotcheck/ct.txt","w")
                        tc.truncate(0)
                        tc.writelines("7.0"+"\n")
                        tc.writelines("7.0"+"\n")
                        tc.writelines("7.3"+"\n")
                        tc.writelines("7.7"+"\n")
                    if(radio_select==1 and kit2_button['bg']=='lawn green'):
                        tc= open("/home/pi/Spotcheck/ct.txt","w")
                        tc.truncate(0)
                        tc.writelines("7.0"+"\n")
                        tc.writelines("7.0"+"\n")
                        tc.writelines("7.8"+"\n")
                        tc.writelines("8"+"\n")
                    if(radio_select==2 and kit2_button['bg']=='lawn green'):
                        tc= open("/home/pi/Spotcheck/ct.txt","w")
                        tc.truncate(0)
                        tc.writelines("7.0"+"\n")
                        tc.writelines("7.0"+"\n")
                        tc.writelines("7.7"+"\n")
                        tc.writelines("7.9"+"\n")

                    messagebox.showinfo("", "Đã lưu xong !")

            def back_click():
                config_click()
            save_button = Button(configmc1_labelframe, bg="yellow", text="Lưu", borderwidth=0, height=3, width=10, command=save_click)
            save_button.place(x=318,y=390)
            back_button = Button(configmc1_labelframe, bg="grey88", text="Trở lại", borderwidth=0, height=3, width=10, command=back_click)
            back_button.place(x=188,y=390)

        def server_click():
            global server_on, ftp_ip , ftp_user, ftp_password, ftp_folder

            configmc2_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
            configmc2_labelframe.place(x=172,y=0)

            server_label = Label(configmc2_labelframe, text='FTP SERVER',font=('bold'), width=61, bg='dodger blue')
            server_label.place(x=3,y=1)

            ip_label = Label(configmc2_labelframe, bg='white', text='Địa chỉ IP', font=('Courier',13,'bold'))
            ip_label.place(x=50,y=157)
            user_label = Label(configmc2_labelframe, bg='white', text='Tên đăng nhập', font=('Courier',13,'bold'))
            user_label.place(x=50,y=207)
            password_label = Label(configmc2_labelframe, bg='white', text='Mật khẩu', font=('Courier',13,'bold'))
            password_label.place(x=50,y=257)
            folder_label = Label(configmc2_labelframe, bg='white', text='Đường dẫn thư mục', font=('Courier',13,'bold'))
            folder_label.place(x=50,y=307)

            ip_entry = Entry(configmc2_labelframe,width=28, font=('Courier',14))
            ip_entry.place(x=253,y=155)
            user_entry = Entry(configmc2_labelframe,width=28, font=('Courier',14))
            user_entry.place(x=253,y=205)
            password_entry = Entry(configmc2_labelframe,width=28, font=('Courier',14))
            password_entry.place(x=253,y=255)
            folder_entry = Entry(configmc2_labelframe,width=28, font=('Courier',14))
            folder_entry.place(x=253,y=305)


            def on_click():
                on_button['bg']='lawn green'
                on_button['fg'] = 'black'
                off_button['bg']='grey88'
                off_button['fg'] = 'grey70'
                ip_entry['state'] = 'normal'
                user_entry['state'] = 'normal'
                password_entry['state'] = 'normal'
                folder_entry['state'] = 'normal'
                ip_entry.delete(0,END)
                ip_entry.insert(0,ftp_ip)
                user_entry.insert(0,ftp_user)
                folder_entry.insert(0,ftp_folder)

            def off_click():
                off_button['bg']='lawn green'
                off_button['fg'] = 'black'
                on_button['bg']='grey88'
                on_button['fg'] = 'grey70'
                ip_entry.delete(0,END)
                user_entry.delete(0,END)
                password_entry.delete(0,END)
                folder_entry.delete(0,END)
                ip_entry['state'] = 'disabled'
                user_entry['state'] = 'disabled'
                password_entry['state'] = 'disabled'
                folder_entry['state'] = 'disabled'


            if(server_on==1):
                on_button = Button(configmc2_labelframe, bg="lawn green", text="Bật", borderwidth=0, height=2, width=7,command=on_click)
                on_button.place(x=302,y=85)
                off_button = Button(configmc2_labelframe, bg="grey88",fg='grey70', text="Tắt", borderwidth=0, height=2, width=7,command=off_click)
                off_button.place(x=220,y=85)
                ip_entry.insert(0,ftp_ip)
                user_entry.insert(0,ftp_user)
                folder_entry.insert(0,ftp_folder)

            else:
                on_button = Button(configmc2_labelframe, bg="grey88", fg='grey70', text="Bật", borderwidth=0, height=2, width=7, command=on_click)
                on_button.place(x=302,y=85)
                off_button = Button(configmc2_labelframe, bg="lawn green", text="Tắt", borderwidth=0, height=2, width=7, command=off_click)
                off_button.place(x=220,y=85)
                ip_entry.delete(0,END)
                user_entry.delete(0,END)
                password_entry.delete(0,END)
                folder_entry.delete(0,END)
                ip_entry['state'] = 'disabled'
                user_entry['state'] = 'disabled'
                password_entry['state'] = 'disabled'
                folder_entry['state'] = 'disabled'


            def save_click():
                msg = messagebox.askquestion("Lưu ", "Bạn có muốn lưu cài đặt ?")
                if(msg=='yes'):
                    if(on_button['bg']=='lawn green'):
                        ip_set = ip_entry.get()
                        user_set = user_entry.get()
                        password_set = password_entry.get()
                        folder_set = folder_entry.get()
                        if(ip_set==''):
                            messagebox.showwarning("","Bạn chưa nhập IP !")
                        elif(user_set==''):
                            messagebox.showwarning("","Bạn chưa nhập Tên đăng nhập !")
                        elif(ip_set==''):
                            messagebox.showwarning("","Bạn chưa nhập Mật khẩu !")
                        elif(folder_set==''):
                            messagebox.showwarning("","Bạn chưa nhập Đường dẫn thư mục !")
                        else:
                            try:
                                ftp = FTP(ip_set, user_set, password_set)
                                ftp.cwd(folder_set)
                                ftp.quit()
                                tc= open("/home/pi/Spotcheck/.server.txt","w")
                                tc.writelines('1\n')
                                tc.writelines(ip_set+"\n")
                                tc.writelines(user_set+"\n")
                                tc.writelines(password_set+"\n")
                                tc.writelines(folder_set+"\n")
                                global server_on, ftp_ip , ftp_user, ftp_password, ftp_folder
                                server_on = 1
                                ftp_ip = ip_set
                                ftp_user = user_set
                                ftp_password = password_set
                                ftp_folder = folder_set
                                messagebox.showinfo("", "Đã lưu xong !")
                            except Exception as e :
                                error = messagebox.showwarning("Không thể kết nối đến Server !",str(e))
                                if(error=='ok'):
                                    pass
                    else:
                        tc= open("/home/pi/Spotcheck/.server.txt","w")
                        tc.writelines('0\n')
                        tc.writelines("\n")
                        tc.writelines("\n")
                        tc.writelines("\n")
                        tc.writelines("\n")
                        server_on = 0
                        messagebox.showinfo("", "Đã lưu xong !")

            def back_click():
                config_click()
            save_button = Button(configmc2_labelframe, bg="yellow", text="Lưu", borderwidth=0, height=3, width=10, command=save_click)
            save_button.place(x=318,y=390)
            back_button = Button(configmc2_labelframe, bg="grey88", text="Trở lại", borderwidth=0, height=3, width=10, command=back_click)
            back_button.place(x=188,y=390)

        ct_button = Button(configmc_labelframe, bg="grey88", text="Chọn kit ly trích", borderwidth=0, height=4, width=15, command=ct_click)
        ct_button.place(x=230,y=150)
        server_button = Button(configmc_labelframe, bg="grey85", text="Server", borderwidth=0, height=4, width=15, command=server_click)
        server_button.place(x=230,y=240)

    home_button = Button(mainscreen_labelframe, bg="dodger blue", activebackground="dodger blue", text="TRANG CHỦ ", fg='white', font=buttonFont, borderwidth=0, height=4, width=20,command=home_click)
    home_button.place(x=1,y=1)
    home_canvas = Canvas(mainscreen_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    home_canvas.place(x=1,y=3)
    setid_button = Button(mainscreen_labelframe, bg="dodger blue", activebackground="dodger blue", text="TỆP\nMẪU XÉT NGHIỆM", fg='white', font=buttonFont, borderwidth=0, height=4, width=20, command=setid_click)
    setid_button.place(x=1,y=81)
    setid_canvas = Canvas(mainscreen_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    setid_canvas.place(x=1,y=83)
    covid19_button = Button(mainscreen_labelframe, bg="dodger blue", activebackground="dodger blue", text="PHÂN TÍCH", fg='white', font=buttonFont, borderwidth=0, height=4, width=20, command=covid19_click)
    covid19_button.place(x=1,y=161)
    covid19_canvas = Canvas(mainscreen_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    covid19_canvas.place(x=1,y=163)
    viewresult_button = Button(mainscreen_labelframe, bg="dodger blue", activebackground="dodger blue", text="XEM KẾT QUẢ", fg='white', font=buttonFont, borderwidth=0, height=4, width=20, command=viewresult_click)
    viewresult_button.place(x=1,y=241)
    viewresult_canvas = Canvas(mainscreen_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    viewresult_canvas.place(x=1,y=243)
    config_button = Button(mainscreen_labelframe, bg="dodger blue", activebackground="dodger blue", text="CÀI ĐẶT", fg='white', font=buttonFont, borderwidth=0, height=4, width=20, command=config_click)
    config_button.place(x=1,y=321)
    config_canvas = Canvas(mainscreen_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    config_canvas.place(x=1,y=323)
    power_button = Button(mainscreen_labelframe, bg="dodger blue", activebackground="dodger blue", text="THOÁT", fg='white', font=buttonFont, borderwidth=0, height=4, width=20, command=power_click)
    power_button.place(x=1,y=401)
    power_canvas = Canvas(mainscreen_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    power_canvas.place(x=1,y=403)

    global covid19clicked
    if(covid19clicked==1):
        covid19_click()
    else:
        home_click()
############################################################# MAIN SCREEN - END ####################################################################

############################################################ VIEW RESULT - START ###################################################################
def result():
    result_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    result_labelframe.place(x=0,y=0)

    result1_labelframe = LabelFrame(result_labelframe, bg='dodger blue', width=795, height=54)
    result1_labelframe.place(x=0,y=424)

    path_label = Label(result_labelframe, bg='red',text=rsfile, font=('Courier',9), width=112)
    path_label.place(x=4,y=4)

    result_img = Image.open(rsfile)
    result_width, result_height = result_img.size
    scale_percent = 81
    width = int(result_width * scale_percent / 100)
    height = int(result_height * scale_percent / 100)
    display_img = result_img.resize((width,height))
    image_select = ImageTk.PhotoImage(display_img)
    result_label = Label(result_labelframe, bg='white',image=image_select)
    result_label.image = image_select
    result_label.place(x=72,y=26)

    def back_click():
        result_labelframe.place_forget()
        mainscreen()
    def open_click():
        global rsfile
        p = rsfile
#         a=0
#         for i in range(len(p)):
#             if(p[i]=='/'):
#                 a=i
#         oldpath = p[:a]
        rsfile = filedialog.askopenfilename(initialdir='/home/pi/Spotcheck Ket Qua', filetypes=[('jpg file','*.jpg')])
        if rsfile is not None:
            if(rsfile[len(rsfile)-3:]=='jpg'):
                result_img = Image.open(rsfile)
                result_width, result_height = result_img.size
                scale_percent = 81
                width = int(result_width * scale_percent / 100)
                height = int(result_height * scale_percent / 100)
                display_img = result_img.resize((width,height))
                image_select = ImageTk.PhotoImage(display_img)

                result_label = Label(result_labelframe, bg='white',image=image_select)
                result_label.image = image_select
                result_label.place(x=72,y=26)

                path_label['text']=rsfile
            else:
                 pass
    back_button = Button(result_labelframe, bg="lavender", text="Trở lại" , height=2, width=8, borderwidth=0, command=back_click)
    back_button.place(x=290,y=428)
    open_button = Button(result_labelframe, bg="lavender", text="Xem tiếp" , height=2, width=8, borderwidth=0, command=open_click)
    open_button.place(x=415,y=428)
############################################################# VIEW RESULT - END ####################################################################

########################################################### SET ID SCREEN - START ##################################################################
def setid():
    setid1_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    setid1_labelframe.place(x=0,y=0)

    setid2_labelframe = LabelFrame(setid1_labelframe, bg='white', width=470, height=160)
    setid2_labelframe.place(x=320,y=5)

    idpos_label = Label(setid2_labelframe, bg='dodger blue', font=("Courier",24,"bold"))
    idpos_label.place(x=1,y=1)

    setidtable_labelframe = LabelFrame(setid1_labelframe,bg='ghost white', width=600, height=307)
    setidtable_labelframe.place(x=10,y=5)

    def idpos_click(n):
        if(idpos_button[n]['bg'] != 'lawn green'):
            for k in range (0,48):
                if(idpos_button[k]['bg'] != 'lawn green' and idpos_button[k]['bg'] != 'grey99'):
                    idpos_button[k]['bg'] = 'lavender'
                else:
                    idpos_button[k]['bg'] = 'lawn green'
            idpos_button[n]['bg'] = 'dodger blue'
        else:
            for k in range (0,48):
                if(idpos_button[k]['bg'] != 'lawn green' and idpos_button[k]['bg'] != 'grey99'):
                    idpos_button[k]['bg'] = 'lavender'
                if(idpos_button[k]['bg'] == 'grey99'):
                    idpos_button[k]['bg'] = 'lawn green'
            idpos_button[n]['bg'] = 'grey99'

        def enter_entry(event):
            try:
                subprocess.Popen(['killall','florence'])
            except:
                pass
            root.attributes('-fullscreen', False)
            subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
            subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)

        def ok_click(event=None):
            if(id_entry.get()==''):
                idpos_button[n]['bg'] = 'lavender'
                idpos_button[n]['text'] = '#'+str(n+1)
                msgbox = messagebox.showwarning(" ","Bạn chưa nhập ID !")
            else:
                idpos_button[n]['text'] = id_entry.get()
                idpos_button[n]['bg'] = 'lawn green'
                try:
                    if(n==45):
                        idpos_click(0)
                    else:
                        idpos_click(n+1)
                except:
                    idpos_click(0)

        id_entry = Entry(setid2_labelframe,width=25, font=('Courier',14))
        if(idpos_button[n]['bg'] == 'grey99'):
            id_entry.insert(0,idpos_button[n]['text'])
        #id_entry.bind("<Button-1>", enter_entry)
        id_entry.bind("<Return>", ok_click)
        id_entry.place(x=50,y=70)
        id_entry.focus_set()

        setid_label = Label(setid2_labelframe, text='Nhập mẫu xét nghiệm', bg='white', font=("Courier",15,"bold"))
        setid_label.place(x=48,y=43)

        if(n<8):
            idpos_label['text'] = str(chr(65+n)) + '1'
        if(n>=8 and n<16):
            idpos_label['text'] = str(chr(65+n-8)) + '2'
        if(n>=16 and n<24):
            idpos_label['text'] = str(chr(65+n-16)) + '3'
        if(n>=24 and n<32):
            idpos_label['text'] = str(chr(65+n-24)) + '4'
        if(n>=32 and n<40):
            idpos_label['text'] = str(chr(65+n-32)) + '5'
        if(n>=40):
            idpos_label['text'] = str(chr(65+n-40)) + '6'

        ok_button = Button(setid2_labelframe, font=('Courier','12','bold'), bg="lavender", text="Xác nhận", height=2, width=8, borderwidth=0, command=ok_click)
        ok_button.place(x=340,y=58)

    idpos_button = list(range(48))
    h=-1
    c=0
    for i in range(0,48):
        h+=1
        if(i%8==0 and i!=0):
            h=0
            c+=1
        idpos_button[i] = Button(setidtable_labelframe, bg='lavender', activebackground="white", justify='left', borderwidth=0, text='#'+str(i+1), width=2, height=2)
        idpos_button[i]['command'] = partial(idpos_click,i)
        idpos_button[i].grid(row=h,column=c,padx=4,pady=4)
        if(i==46):
            idpos_button[i]['state']='disabled'
            idpos_button[i]['bg']= 'green'
            idpos_button[i]['text']= 'N'
        if(i==47):
            idpos_button[i]['state']='disabled'
            idpos_button[i]['bg']= 'red'
            idpos_button[i]['text']= 'P'

    def cancel_click():
        msg = messagebox.askquestion("Hủy", "Bạn muốn hủy mà không lưu lại tệp ?")
        if(msg=="yes"):
            setid1_labelframe.place_forget()
            mainscreen()

    def save_click():
        workbook = load_workbook("/home/pi/Spotcheck/template.xlsm", keep_vba = True)
        sheet = workbook.active
        # for i in range(0,48):
        #     #pos = "C"+str(i+3)
        #     if(i<8):
        #         pos = 'B'+ str(i+2)
        #     if(i>=8 and i<16):
        #         pos = 'C'+ str(i-6)
        #     if(i>=16 and i<24):
        #         pos = 'D'+ str(i-14)
        #     if(i>=24 and i<32):
        #         pos = 'E'+ str(i-22)
        #     if(i>=32 and i<40):
        #         pos = 'F'+ str(i-30)
        #     if(i>=40):
        #         pos = 'G'+str(i-38)

        for i in range(0,48):
            pos = "B" + str(i+12)
            if(idpos_button[i]['bg']=='lawn green' or idpos_button[i]['bg']=='grey99'):
                sheet[pos] = idpos_button[i]['text']
            else:
                sheet[pos] = 'N/A'

        sheet['B58']='NEGC'
        sheet['B59']='POSC'

        msg = messagebox.askquestion("Lưu ", "Bạn có muốn lưu tệp ?")
        if(msg=='yes'):
            f = filedialog.asksaveasfilename(initialdir='/home/pi/Desktop/Spotcheck ID/',defaultextension='.xlsx')
            if f is not None:
                d=0
                for i in range(len(f)):
                    if(f[i]=='/'):
                        d=i+1
                filename = f[d:(len(f)-5)]
                print(filename)
                if(len(filename)<=30):
                    workbook.save(f)
                    try:
                        subprocess.Popen(['killall','florence'])
                    except:
                        pass
                    root.attributes('-fullscreen', True)

                    msg = messagebox.askquestion(' ','Đã lưu xong!\nBạn có muốn tạo tệp mới ?')
                    if(msg=='yes'):
                        setid()
                    else:
                        setid1_labelframe.place_forget()
                        mainscreen()

                else:
                    messagebox.showerror("Lỗi", "Tên tệp không vượt quá 30 ký tự !")

    def load_click():
        idfile = filedialog.askopenfilename(initialdir='/home/pi/Desktop/Spotcheck ID', filetypes=[('Excel file','*.xlsm *.xlsx *.xls')])
        if idfile is not None:
            if(idfile[len(idfile)-4:]=='xlsx' or idfile[len(idfile)-4:]=='xlsm' or idfile[len(idfile)-3:]=='xls'):
                workbook = openpyxl.load_workbook(idfile)
                sheet = workbook.active
                idfile_list = list(range(48))

                # for i in range(0,48):
                #     if(i<8):
                #         pos = 'B'+ str(i+2)
                #     if(i>=8 and i<16):
                #         pos = 'C'+ str(i-6)
                #     if(i>=16 and i<24):
                #         pos = 'D'+ str(i-14)
                #     if(i>=24 and i<32):
                #         pos = 'E'+ str(i-22)
                #     if(i>=32 and i<40):
                #         pos = 'F'+ str(i-30)
                #     if(i>=40):
                #         pos = 'G'+str(i-38)

                for i in range(0,48):
                    pos = 'B' + str(i+12)
                    idfile_list[i] = sheet[pos].value
                    idpos_button[i]['text'] = idfile_list[i]
                    if(idpos_button[i]['text']!='N/A'):
                        idpos_button[i]['bg']='lawn green'
                    if(i==46):
                        idpos_button[i]['bg']= 'green'
                    if(i==47):
                        idpos_button[i]['bg']= 'red'
        else:
            pass
    def keyboard_click():
        if(keyboard_button['bg']=='grey85'):
            keyboard_button['bg']='lawn green'
            try:
                subprocess.Popen(['killall','florence'])
            except:
                pass
            root.attributes('-fullscreen', False)
            subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
            subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
        else:
            keyboard_button['bg']='grey85'
            try:
                subprocess.Popen(['killall','florence'])
            except:
                pass
            root.attributes('-fullscreen', True)

    idpos_click(0)

    cancel_button = Button(setid1_labelframe, font=('Courier','12','bold'), bg="lavender", text="Hủy" , height=3, width=11, borderwidth=0, command=cancel_click)
    cancel_button.place(x=653,y=170)
    save_button = Button(setid1_labelframe, activebackground="gold", font=('Courier','12','bold'), bg="yellow", text="Lưu", height=3, width=11, borderwidth=0, command=save_click)
    save_button.place(x=487,y=170)
    load_button = Button(setid1_labelframe, font=('Courier','12','bold'), bg="lavender", text="Chỉnh sửa\ntệp sẵn có", height=3, width=11, borderwidth=0, command=load_click)
    load_button.place(x=320,y=170)
    keyboard_button = Button(setid1_labelframe, font=('Courier','10','bold'), bg="grey85", text="Bàn phím", height=3, width=7, borderwidth=0, command=keyboard_click)
    keyboard_button.place(x=706,y=374)
############################################################ SET ID SCREEN - END ###################################################################

###################################################### SET TEMPERATURES SCREEN - START #############################################################
def settemp():
    if(covid19clicked==1):
        fr = open("/home/pi/Spotcheck/covid19saved.txt","r")
    if(tbclicked==1):
        fr = open("/home/pi/Spotcheck/tbsaved.txt","r")
    if(spotcheckclicked==1):
        fr = open("/home/pi/Spotcheck/scsaved.txt","r")
    if(shrimpclicked==1):
        fr = open("/home/pi/Spotcheck/shrimpsaved.txt","r")
    t1 = fr.readline()[3:5]
    t2 = fr.readline()[3:5]
    t3 = fr.readline()[3:5]
#     thr1 = fr.readline()[5:9]
#     thr2 = fr.readline()[5:9]
#     thr3l = fr.readline()[6:10]
#     thr3h = fr.readline()[6:10]
    global samples
    samples=0
    settemp_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    settemp_labelframe.place(x=0,y=0)
    settemptop_labelframe = LabelFrame(settemp_labelframe, bg='white', width=798, height=350)
    settemptop_labelframe.place(x=0,y=52)
    keypad_labelframe = LabelFrame(settemptop_labelframe, bg='white', width=285, height=323)
    keypad_labelframe.place(x=501,y=11)
    title_labelframe = LabelFrame(settemp_labelframe, bg='dodger blue', width=798, height=50)
    title_labelframe.place(x=0,y=0)
    settemp_label = Label(settemp_labelframe, bg='dodger blue', fg='black', text='CÀI ĐẶT NHIỆT ĐỘ', font=("Courier",17,'bold'), width=20, height=1 )
    settemp_label.place(x=260,y=12)

    def numpad_click(btn):
        text = "%s" % btn
        if (text!="Xóa" and text!='Mặc định'):
            if(entry_num==1):
                t1_entry.insert(END, text)
            if(entry_num==2):
                t2_entry.insert(END, text)
            if(entry_num==3):
                t3_entry.insert(END, text)
#             if(entry_num==4):
#                 thr1_entry.insert(END, text)
#             if(entry_num==5):
#                 thr2_entry.insert(END, text)
#             if(entry_num==6):
#                 thr3l_entry.insert(END, text)
#             if(entry_num==7):
#                 thr3h_entry.insert(END, text)
        if text == 'Xóa':
            if(entry_num==1):
                t1_entry.delete(0, END)
            if(entry_num==2):
                t2_entry.delete(0, END)
            if(entry_num==3):
                t3_entry.delete(0, END)
#             if(entry_num==4):
#                 thr1_entry.delete(0, END)
#             if(entry_num==5):
#                 thr2_entry.delete(0, END)
#             if(entry_num==6):
#                 thr3l_entry.delete(0, END)
#             if(entry_num==7):
#                 thr3h_entry.delete(0, END)
        if text == 'Mặc định':
            if(entry_num==1):
                t1_entry.delete(0, END)
                t1_entry.insert(END, t1)
            if(entry_num==2):
                t2_entry.delete(0, END)
                t2_entry.insert(END, t2)
            if(entry_num==3):
                t3_entry.delete(0, END)
                t3_entry.insert(END, t3)
#             if(entry_num==4):
#                 thr1_entry.delete(0, END)
#                 thr1_entry.insert(END, 25)
#             if(entry_num==5):
#                 thr2_entry.delete(0, END)
#                 thr2_entry.insert(END, 25)
#             if(entry_num==6):
#                 thr3l_entry.delete(0, END)
#                 thr3l_entry.insert(END, 25)
#             if(entry_num==7):
#                 thr3h_entry.delete(0, END)
#                 thr3h_entry.insert(END, 25)

    def numpad():
        global numpad_labelframe
        numpad_labelframe = LabelFrame(keypad_labelframe, bg="white", width=385, height=395)
        numpad_labelframe.place(x=2,y=1)
        button_list = ['7',     '8',      '9',
                       '4',     '5',      '6',
                       '1',     '2',      '3',
                       '0',     'Xóa', 'Mặc định']
        r = 1
        c = 0
        n = 0
        btn = list(range(len(button_list)))
        for label in button_list:
            cmd = partial(numpad_click, label)
            btn[n] = Button(numpad_labelframe, text=label, font=font.Font(family='Helvetica', size=10, weight='bold'), width=9, height=4, command=cmd)
            btn[n].grid(row=r, column=c, padx=0, pady=0)
            n += 1
            c += 1
            if (c == 3):
                c = 0
                r += 1

    temp_labelframe = LabelFrame(settemptop_labelframe, text='NHIỆT ĐỘ', bg='white', width=490, height=180)
    temp_labelframe.place(x=3,y=2)
#     thres_labelframe = LabelFrame(settemptop_labelframe, text='THRESHOLD', bg='white', width=490, height=149)
#     thres_labelframe.place(x=3,y=185)

    cir_img = Image.open('/home/pi/Spotcheck/cir.png')
    cir_width, cir_height = cir_img.size
    scale_percent = 14
    width = int(cir_width * scale_percent / 100)
    height = int(cir_height * scale_percent / 100)
    display_img = cir_img.resize((width,height))
    image_select = ImageTk.PhotoImage(display_img)
    t1cir_label = Label(temp_labelframe, bg='white', image=image_select)
    t1cir_label.image = image_select
    t1cir_label.place(x=5,y=5)
    t2cir_label = Label(temp_labelframe, bg='white', image=image_select)
    t2cir_label.image = image_select
    t2cir_label.place(x=170,y=5)
    t3cir_label = Label(temp_labelframe, bg='white', image=image_select)
    t3cir_label.image = image_select
    t3cir_label.place(x=335,y=5)
#     graycir_img = Image.open('graycir.png')
#     graycir_width, cir_height = cir_img.size
#     scale_percent = 16
#     width = int(cir_width * scale_percent / 100)
#     height = int(cir_height * scale_percent / 100)
#     display_img = graycir_img.resize((width,height))
#     image_select = ImageTk.PhotoImage(display_img)
#     t4cir_label = Label(settemptop_labelframe, bg='white', image=image_select)
#     t4cir_label.image = image_select
#     t4cir_label.place(x=275,y=175)

    def entryt1_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 1
        numpad()
    def entryt2_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 2
        numpad()
    def entryt3_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 3
        numpad()
#     def entrythr1_click(event):
#         global numpad_labelframe
#         global entry_num
#         entry_num = 4
#         numpad()
#     def entrythr2_click(event):
#         global numpad_labelframe
#         global entry_num
#         entry_num = 5
#         numpad()
#     def entrythr3l_click(event):
#         global numpad_labelframe
#         global entry_num
#         entry_num = 6
#         numpad()
#     def entrythr3h_click(event):
#         global numpad_labelframe
#         global entry_num
#         entry_num = 7
#         numpad()

    t1_label = Label(temp_labelframe, bg='white', text='T1', fg='black', font=("Courier",20,"bold"))
    t1_label.place(x=15, y=14)
    t1oc_label = Label(temp_labelframe, bg='white', text=chr(176)+'C', fg='red', font=("Courier", 11,"bold"))
    t1oc_label.place(x=107, y=55)
    t1_entry = Entry(temp_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",36,"bold"))
    t1_entry.place(x=47,y=50)
    t1_entry.bind('<Button-1>', entryt1_click)
    t1_entry.insert(0,t1)

    t2_label = Label(temp_labelframe, bg='white', text='T2', fg='black', font=("Courier",20,"bold"))
    t2_label.place(x=180, y=14)
    t2oc_label = Label(temp_labelframe, bg='white', text=chr(176)+'C', fg='red', font=("Courier", 11,"bold"))
    t2oc_label.place(x=272, y=55)
    t2_entry = Entry(temp_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",36,"bold"))
    t2_entry.place(x=212,y=50)
    t2_entry.bind('<Button-1>', entryt2_click)
    t2_entry.insert(0,t2)

    t3_label = Label(temp_labelframe, bg='white', text='T3', fg='black', font=("Courier",20,"bold"))
    t3_label.place(x=345, y=14)
    t3oc_label = Label(temp_labelframe, bg='white', text=chr(176)+'C', fg='red', font=("Courier", 11,"bold"))
    t3oc_label.place(x=437, y=55)
    t3_entry = Entry(temp_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",36,"bold"))
    t3_entry.place(x=377,y=50)
    t3_entry.bind('<Button-1>', entryt3_click)
    t3_entry.insert(0,t3)

#     thr1_label = Label(thres_labelframe, bg='white', text='T1: ', fg='black', font=("Courier",24,"bold"))
#     thr1_label.place(x=60, y=7)
#     thr1_entry = Entry(thres_labelframe, width=4, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",25,"bold"))
#     thr1_entry.place(x=118,y=7)
#     thr1_entry.bind('<Button-1>', entrythr1_click)
#     thr1_entry.insert(0,thr1)
#
#     thr2_label = Label(thres_labelframe, bg='white', text='T2: ', fg='black', font=("Courier",24,"bold"))
#     thr2_label.place(x=60, y=71)
#     thr2_entry = Entry(thres_labelframe, width=4, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",25,"bold"))
#     thr2_entry.place(x=118,y=71)
#     thr2_entry.bind('<Button-1>', entrythr2_click)
#     thr2_entry.insert(0,thr2)
#
#     thr3l_label = Label(thres_labelframe, bg='white', text='T3-L: ', fg='black', font=("Courier",24,"bold"))
#     thr3l_label.place(x=235, y=7)
#     thr3l_entry = Entry(thres_labelframe, width=4, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",25,"bold"))
#     thr3l_entry.place(x=330,y=7)
#     thr3l_entry.bind('<Button-1>', entrythr3l_click)
#     thr3l_entry.insert(0,thr3l)
#
#     thr3h_label = Label(thres_labelframe, bg='white', text='T3-H: ', fg='black', font=("Courier",24,"bold"))
#     thr3h_label.place(x=235, y=71)
#     thr3h_entry = Entry(thres_labelframe, width=4, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",25,"bold"))
#     thr3h_entry.place(x=330,y=71)
#     thr3h_entry.bind('<Button-1>', entrythr3h_click)
#     thr3h_entry.insert(0,thr3h)

#     t4_label = Label(settemptop_labelframe, bg = 'white', text='T4', fg='grey67', font=("Courier",20,"bold"))
#     t4_label.place(x=286, y=185)

    def back_click():
        settemp_labelframe.place_forget()
        mainscreen()
    def thread():
        th1 = Thread(target = next_click)
        th1.start()
    def next_click():
        try:
            camera.close()
        except:
            pass
        settemp_labelframe.place_forget()
        global t1_set, t2_set, t3_set
#         global thr1_set, thr2_set, thr3l_set, thr3h_set
        t1_set = t1_entry.get()[0:2]
        t2_set = t2_entry.get()[0:2]
        t3_set = t3_entry.get()[0:2]
#         thr1_set = thr1_entry.get()[0:4]
#         thr2_set = thr2_entry.get()[0:4]
#         thr3l_set = thr3l_entry.get()[0:4]
#         thr3h_set = thr3h_entry.get()[0:4]

        global path5
        if os.path.exists(path5+"/nhiet-do.txt"):
            fc= open(path5+"/nhietdo.txt","w")
            fc.truncate(0)
            fc.writelines("T1="+t1_entry.get()[0:2]+"\n")
            fc.writelines("T2="+t2_entry.get()[0:2]+"\n")
            fc.writelines("T3="+t3_entry.get()[0:2]+"\n")
        else:
            fc= open(path5+"/nhietdo.txt","w+")
            fc.writelines("T1="+t1_entry.get()[0:2]+"\n")
            fc.writelines("T2="+t2_entry.get()[0:2]+"\n")
            fc.writelines("T3="+t3_entry.get()[0:2]+"\n")
        scanposition()
    def save_click():
        msg = messagebox.askquestion("Lưu chương trình nhiệt", "Bạn có muốn lưu nhiệt độ ?")
        if(msg=='yes'):
            messagebox.showinfo("","Đã lưu xong !")
            if(covid19clicked==1):
                fw = open("/home/pi/Spotcheck/covid19saved.txt","w")
            if(tbclicked==1):
                fw = open("/home/pi/Spotcheck/tbsaved.txt","w")
            if(spotcheckclicked==1):
                fw = open("/home/pi/Spotcheck/scsaved.txt","w")
            if(shrimpclicked==1):
                fw = open("/home/pi/Spotcheck/shrimpsaved.txt","w")
            fw.truncate(0)
            fw.writelines("T1="+t1_entry.get()[0:2]+"\n")
            fw.writelines("T2="+t2_entry.get()[0:2]+"\n")
            fw.writelines("T3="+t3_entry.get()[0:2]+"\n")
#         if(len(thr1_entry.get())<=2):
#             fw.writelines("THR1="+thr1_entry.get()[0:2]+".0"+"\n")
#         else:
#             fw.writelines("THR1="+thr1_entry.get()[0:4]+"\n")
#         if(len(thr2_entry.get())<=2):
#             fw.writelines("THR1="+thr2_entry.get()[0:2]+".0"+"\n")
#         else:
#             fw.writelines("THR2="+thr2_entry.get()[0:4]+"\n")
#         if(len(thr3l_entry.get())<=2):
#             fw.writelines("THR1="+thr3l_entry.get()[0:2]+".0"+"\n")
#         else:
#             fw.writelines("THR3L="+thr3l_entry.get()[0:4]+"\n")
#         if(len(thr3h_entry.get())<=2):
#             fw.writelines("THR1="+thr3h_entry.get()[0:2]+".0"+"\n")
#         else:
#             fw.writelines("THR3H="+thr3h_entry.get()[0:4]+"\n")

    back_button = Button(settemp_labelframe, font=('Courier','12','bold'), bg="lavender", text="Trở lại" , height=3, width=11, borderwidth=0, command=back_click)
    back_button.place(x=14,y=406)
    next_button = Button(settemp_labelframe, font=('Courier','12','bold'), bg="lavender", text="Tiếp theo", height=3, width=11, borderwidth=0, command=thread)
    next_button.place(x=647,y=406)
    save_button = Button(settemp_labelframe, activebackground="gold", font=('Courier','12','bold'), bg="yellow", text="Lưu", height=3, width=11, borderwidth=0,command=save_click)
    save_button.place(x=332,y=406)
####################################################### SET TEMPERATURES SCREEN - END ##############################################################

######################################################### SAMPLES POSITION - START #################################################################
def scanposition():
    print(thr1_set)
    print(thr2_set)
    print(thr3l_set)
    print(thr3h_set)
    global path0
    global path1
    global path2
    global path3
    global path4
    global path5

    global ser
    ser.flushInput()
    ser.flushOutput()
    global scanpostion_labelframe
    scanposition_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    scanposition_labelframe.place(x=0,y=0)
    title_labelframe = LabelFrame(scanposition_labelframe, bg='dodger blue', width=798, height=50)
    title_labelframe.place(x=0,y=0)
    scanposition_label = Label(scanposition_labelframe, bg='dodger blue', text='XÁC ĐỊNH VỊ TRÍ MẪU', font=("Courier",17,'bold'), width=20, height=1 )
    scanposition_label.place(x=258,y=12)

    scan_img = Image.open('/home/pi/Spotcheck/scan.png')
    scan_width, scan_height = scan_img.size
    scale_percent = 100
    width = int(scan_width * scale_percent / 100)
    height = int(scan_height * scale_percent / 100)
    display_img = scan_img.resize((width,height))
    image_select = ImageTk.PhotoImage(display_img)
    scan_label = Label(scanposition_labelframe, bg='white',image=image_select)
    scan_label.image = image_select
    scan_label.place(x=270,y=80)

    s = ttk.Style()
    s.theme_use('clam')
    s.configure("green.Horizontal.TProgressbar", foreground='green', background='green')
    scanposition_progressbar = ttk.Progressbar(root, orient = HORIZONTAL, style="green.Horizontal.TProgressbar", length = 200, mode = 'determinate')
    scanposition_progressbar.place(x=299,y=408)
    root.update_idletasks()

    def back_click():
        try:
            camera.close()
        except Exception:
            pass
        global wait
        wait = 0
        mainscreen()

    back_button = Button(scanposition_labelframe, font=("Courier",12,'bold'), bg="lavender", text="Trở lại" , height=3, width=11, borderwidth=0, command=back_click)
    back_button.place(x=14,y=406)
    process_label = Label(scanposition_labelframe, text='Đang xử lý ...', bg='white', font=("Courier",13))
    process_label.place(x=330,y=440)

    send_data = 'P'
    ser.write(send_data.encode())

    if(ser.in_waiting>0):
        receive_data = ser.readline().decode('utf-8').rstrip()
        print("Data received:", receive_data)
        scanposition_progressbar['value'] = 5
        root.update_idletasks()
        if(receive_data=='C'):
            global wait
            wait = 1
            scanposition_progressbar['value'] = 20
            root.update_idletasks()

    while(wait!=1):
        scanposition_progressbar['value'] = 2
        root.update_idletasks()
        if(ser.in_waiting>0):
            receive_data = ser.readline().decode('utf-8').rstrip()
            print("Data received:", receive_data)
            scanposition_progressbar['value'] = 10
            root.update_idletasks()
            if(receive_data=='C'):
                scanposition_progressbar['value'] = 20
                root.update_idletasks()
                wait = 1
                break;
    while(wait==1):
        try:
            camera_capture(path4 + "/mau.jpg")
        except Exception as e :
            error = messagebox.askquestion("Lỗi: "+ str(e), "Bạn có muốn thoát chương trình ?", icon = "error")
            if(error=='yes'):
                root.destroy()

        image = cv2.imread(path4 + "/mau.jpg")
        blur_img = cv2.fastNlMeansDenoisingColored(image.copy(),None,15,15,7,21)
        gray_img = cv2.cvtColor(blur_img, cv2.COLOR_BGR2GRAY)
        thresh, binary_img = cv2.threshold(gray_img.copy(), 40, maxval=255, type=cv2.THRESH_BINARY)
        contours, hierarchy = cv2.findContours(binary_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        print("Number of contours: " + str(len(contours)))

        contours.sort(key=lambda data:sorting_xy(data))

        contour_img = np.zeros_like(gray_img)
        bourect0 = cv2.boundingRect(contours[0])
        bourect47 = cv2.boundingRect(contours[len(contours)-1])
        global start_point
        start_point = (bourect0[0]-9, bourect0[1]-9)
        global end_point
        end_point = (bourect47[0]+bourect47[2]+9, bourect47[1]+bourect47[3]+9)
        print('Start point:', start_point)
        print('End point:', end_point)
        fw= open("/home/pi/Spotcheck/coordinates2.txt",'w')
        fw.writelines("Start Point: " + str(start_point) + "\n")
        fw.writelines("End Point: " + str(end_point))

        scanposition_progressbar['value'] = 35
        root.update_idletasks()

        global pos_result
        pos_result, pos_image = process_image(path4 + "/mau.jpg")
        #pos_result, pos_image = process_image("/home/pi/Desktop/mau.jpg")
        scanposition_progressbar['value'] = 60
        root.update_idletasks()
        sleep(1)

        output = path4 + "/xu-ly-mau.jpg"
        cv2.imwrite(output, pos_image)
        scanposition_progressbar['value'] = 90
        root.update_idletasks()
        sleep(1)

        scanresult_labelframe = LabelFrame(scanposition_labelframe, bg='ghost white', width=528,height = 307)
        scanresult_labelframe.place(x=248,y=60)

        label = list(range(48))
        global id_list
        def result_table(range_a, range_b, row_value):
            global samples
            j=-1
            for i in range(range_a, range_b):
                j+=1
                if(i<6):
                    t='A'+ str(i+1)
                if(i>=6 and i<12):
                    t='B'+ str(i-5)
                if(i>=12 and i<18):
                    t='C'+ str(i-11)
                if(i>=18 and i<24):
                    t='D'+ str(i-17)
                if(i>=24 and i<30):
                    t='E'+ str(i-23)
                if(i>=30 and i<36):
                    t='F'+ str(i-29)
                if(i>=36 and i<42):
                    t='G'+ str(i-35)
                if(i>=42):
                    t='H'+ str(i-41)
                if(id_list[i]=='N/A'):
                    label[i] = Label(scanresult_labelframe, bg='white', text=t, width=5, height=2)
                    label[i].grid(row=row_value,column=j,padx=3,pady=3)
                else:
                    if(pos_result[i]<=8):
                        label[i] = Label(scanresult_labelframe, bg='gainsboro', text=t, width=5, height=2)
                        label[i].grid(row=row_value,column=j,padx=3,pady=3)
                    else:
                        label[i] = Label(scanresult_labelframe, bg='OliveDrab1', text=t, width=5, height=2)
                        label[i].grid(row=row_value,column=j,padx=3,pady=3)
                        samples += 1
        scanposition_progressbar['value'] = 100
        root.update_idletasks()

        result_table(0,6,0)
        result_table(6,12,1)
        result_table(12,18,2)
        result_table(18,24,3)
        result_table(24,30,4)
        result_table(30,36,5)
        result_table(36,42,6)
        result_table(42,48,7)
        global samples
        samplenum_label = Label(scanposition_labelframe, text='Số mẫu hiện tại: ' + str(samples), fg='dodger blue', bg='white', font=("Courier",13,'bold'))
        samplenum_label.place(x=300,y=432)
        scan_label.place_forget()
        scanposition_progressbar.place_forget()
        process_label.place_forget()
        wait = 0
        samples = 0
        def thread():
            th1 = Thread(target = next_click)
            th1.start()
        def next_click():
            global createclicked
            createclicked = 0
            scanposition_labelframe.place_forget()
            analysis()
        next_button = Button(scanposition_labelframe, font=("Courier",12,'bold'), bg="lavender", text="Tiếp theo", height=3, width=11, borderwidth=0,command=thread)
        next_button.place(x=647,y=406)

########################################################## SAMPLES POSITION - END ##################################################################

######################################################### SAMPLES ANALYSIS - START #################################################################
def analysis():
    global ser
    ser.flushInput()
    ser.flushOutput()

    global analysis_labelframe
    analysis_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    analysis_labelframe.place(x=0,y=0)
    title_labelframe = LabelFrame(analysis_labelframe, bg='dodger blue', width=798, height=50)
    title_labelframe.place(x=0,y=0)
    analysis_label = Label(analysis_labelframe, bg='dodger blue', text='PHÂN TÍCH MẪU', font=("Courier",17,'bold'), width=20, height=1 )
    analysis_label.place(x=261,y=12)
    t_labelframe = LabelFrame(analysis_labelframe, bg='white', width=798, height=298)
    t_labelframe.place(x=0,y=70)

    #t1_labelframe = LabelFrame(t_labelframe, bg='white',text="T1:"+'RT'+chr(176)+'C' , font=("Courier",13,'bold'), width=197, height=290)
    t1_labelframe = LabelFrame(t_labelframe, bg='white',text="T1", font=("Courier",13,'bold'), width=197, height=290)
    t1_labelframe.place(x=0,y=2)
    #t2_labelframe = LabelFrame(t_labelframe, bg='white',text="T2"+t2_set+chr(176)+'C' , font=("Courier",13,'bold'), width=197, height=290)
    t2_labelframe = LabelFrame(t_labelframe, bg='white',text="T2", font=("Courier",13,'bold'), width=197, height=290)
    t2_labelframe.place(x=199,y=2)
    #t3_labelframe = LabelFrame(t_labelframe, bg='white',text="T3"+t3_set+chr(176)+'C' , font=("Courier",13,'bold'), width=197, height=290)
    t3_labelframe = LabelFrame(t_labelframe, bg='white',text="T3", font=("Courier",13,'bold'), width=197, height=290)
    t3_labelframe.place(x=398,y=2)
    t4_labelframe = LabelFrame(t_labelframe, bg='white smoke',text="T4", width=197, height=290)
    t4_labelframe.place(x=597,y=2)
    t1wait_label = Label(t1_labelframe, text='...', fg='grey36', bg='white', font=("Courier",40,'bold'))
    t1wait_label.place(x=46,y=110)
    t2wait_label = Label(t2_labelframe, text='...', fg='grey36', bg='white', font=("Courier",40,'bold'))
    t2wait_label.place(x=46,y=110)
    t3wait_label = Label(t3_labelframe, text='...', fg='grey36', bg='white', font=("Courier",40,'bold'))
    t3wait_label.place(x=46,y=110)
    temp_label = Label(analysis_labelframe, bg='white', fg='grey36', font=("Courier",20,'bold'))
    temp_label.place(x=65,y=389)

    def stop_click():
        global ser
        msgbox = messagebox.askquestion('Dừng xử lý','Bạn có muốn dừng quá trình xử lý ?', icon = 'question')
        if(msgbox=='yes'):
            send_data ='S'
            ser.write(send_data.encode())
            try:
                camera.close()
            except:
                pass
            analysis_labelframe.place_forget()
            global wait
            wait = 0
            mainscreen()

    # def pause_click():
#         try:
#             camera.close()
#         except:
#             pass
#         global ser
#         if(pause_button['text']=='Pause'):
#             send_data ='P'
#             ser.write(send_data.encode())
#             pause_button['text']= 'Continue'
#         else:
#             send_data ='R'
#             ser.write(send_data.encode())
#             pause_button['text']= 'Pause'

    # pause_button = Button(analysis_labelframe, bg="lavender", font=("Courier",12,'bold'), text="Pause" , height=3, width=10, borderwidth=0, command=pause_click)
#     pause_button.place(x=450,y=390)
    stop_button = Button(analysis_labelframe, bg="red", font=("Courier",12,'bold'), text="Dừng", height=3, width=9, borderwidth=0, command=stop_click)
    stop_button.place(x=600,y=390)
    root.update()

    send_data = "t"+ t1_set + "," + t2_set + "," + t3_set + "z"
    ser.write(send_data.encode())
    print("Data send: ", send_data)
    #t0 = time.time()
    sleep(2)

    global wait
    if(ser.in_waiting>0):
        receive_data = ser.readline().decode('utf-8').rstrip()
        print("Data received:", receive_data)
        if(receive_data=='Y'):
            autoprocess_label = Label(analysis_labelframe, bg='white', text="Chương trình đang xử lý...", fg='blue', font=("Courier",12,'bold'))
            autoprocess_label.place(x=65,y=438)
            wait = 1
    while(wait!=1):
        if(ser.in_waiting>0):
            receive_data = ser.readline().decode('utf-8').rstrip()
            print("Data received:", receive_data)
            if(receive_data=='Y'):
                autoprocess_label = Label(analysis_labelframe, bg='white', text="Chương trình đang xử lý...", fg='blue', font=("Courier",12,'bold'))
                autoprocess_label.place(x=65,y=438)
                wait = 1
                break

    global id_list
    while(wait==1):
        if(ser.in_waiting>0):
            global t1_run, t2_run, t3_run
            receive_data = ser.readline().decode('utf-8',errors='ignore').rstrip()
            #print("Data received:", receive_data)
            if(receive_data!='C1' and receive_data!='C2' and receive_data!='C3'):
                print("Data received:", receive_data)
                temp_label['text'] = 'Nhiệt độ: '+ receive_data + chr(176)+'C'
                root.update()

            if(receive_data=='C1'):
                t1_run=1
                t2_run=0
                t3_run=0
                print("Data received:", receive_data)
                t1wait_label.place_forget()
                t1_labelframe['bg'] = atk.DEFAULT_COLOR
                t1_labelframe['fg'] = 'lawn green'
                t_progressbar = atk.RadialProgressbar(t1_labelframe, fg='cyan')
                t_progressbar.place(x=47,y=70)
                t_progressbar.start()
                tprocess_label = Label(t1_labelframe, bg=atk.DEFAULT_COLOR, fg='white smoke', text='Đang phân tích...', font=("Courier",9,'bold'))
                tprocess_label.place(x=38,y=112)

                global path1
                camera_capture(path1 + "/T1.jpg")

                send_data = 'C'
                ser.write(send_data.encode())
                print('Capture done!')

                global start_point
                global end_point
                #t1_result, t1_image= process_image(path1 + "/T1.jpg", start_point, end_point)
                t1_result, t1_image= process_image(path1 + "/T1.jpg")

                global path2
                output = path2 + "/T1.jpg"
                cv2.imwrite(output, t1_image)

                t1_analysis = Image.open(output)
                t1_crop = t1_analysis.crop((x1-13, y1-13, x2+13, y2+13))
                #t1_crop = t1_analysis.crop((280-7, 81-7, 498+7, 376+7))
                crop_width, crop_height = t1_crop.size
                scale_percent = 75
                width = int(crop_width * scale_percent / 100)
                height = int(crop_height * scale_percent / 100)
                display_img = t1_crop.resize((width,height))
                t1_display = ImageTk.PhotoImage(display_img)
                t1_label = Label(t1_labelframe, image=t1_display)
                t1_label.image = t1_display
                t1_label.place(x=0,y=1)
                root.update()

                workbook = Workbook()
                sheet = workbook.active

                sheet["A2"] = "A"
                sheet["A3"] = "B"
                sheet["A4"] = "C"
                sheet["A5"] = "D"
                sheet["A6"] = "E"
                sheet["A7"] = "F"
                sheet["A8"] = "G"
                sheet["A9"] = "H"
                sheet["B1"] = "1"
                sheet["C1"] = "2"
                sheet["D1"] = "3"
                sheet["E1"] = "4"
                sheet["F1"] = "5"
                sheet["G1"] = "6"
                for i in range(0,48):
                    if(i<6):
                        pos = str(chr(65+i+1)) + "2"
                    if(i>=6 and i<12):
                        pos = str(chr(65+i-5)) + "3"
                    if(i>=12 and i<18):
                        pos = str(chr(65+i-11)) + "4"
                    if(i>=18 and i<24):
                        pos = str(chr(65+i-17)) + "5"
                    if(i>=24 and i<30):
                        pos = str(chr(65+i-23)) + "6"
                    if(i>=30 and i<36):
                        pos = str(chr(65+i-29)) + "7"
                    if(i>=36 and i<42):
                        pos = str(chr(65+i-35)) + "8"
                    if(i>=42):
                        pos = str(chr(65+i-41)) + "9"

                    if(id_list[i]=='N/A'):
                        sheet[pos] = 'N/A'
                    else:
                        sheet[pos] = t1_result[i]

                global path3
                workbook.save(path3+"/T1.xlsx")

            if(receive_data=='C2'):
                t1_run=0
                t2_run=1
                t3_run=0
                print("Data received:", receive_data)
                t2wait_label.place_forget()
                t2_labelframe['bg'] = atk.DEFAULT_COLOR
                t2_labelframe['fg'] = 'lawn green'
                t_progressbar = atk.RadialProgressbar(t2_labelframe, fg='cyan')
                t_progressbar.place(x=47,y=70)
                t_progressbar.start()
                tprocess_label = Label(t2_labelframe, bg=atk.DEFAULT_COLOR, fg='white smoke', text='Đang phân tích...', font=("Courier",9,'bold'))
                tprocess_label.place(x=38,y=112)

                camera_capture(path1 + "/T2.jpg")

                send_data = 'C'
                ser.write(send_data.encode())
                print('Capture done!')
                #t2_result, t2_image = process_image(path1 + "/T2.jpg", start_point, end_point)
                t2_result, t2_image= process_image(path1 + "/T2.jpg")

                output = path2 + "/T2.jpg"
                cv2.imwrite(output, t2_image)
                t2_analysis = Image.open(output)
                t2_crop = t2_analysis.crop((x1-13, y1-13, x2+13, y2+13))
                #t2_crop = t2_analysis.crop((280-7, 81-7, 498+7, 376+7))
                crop_width, crop_height = t2_crop.size
                scale_percent = 75
                width = int(crop_width * scale_percent / 100)
                height = int(crop_height * scale_percent / 100)
                display_img = t2_crop.resize((width,height))
                t2_display = ImageTk.PhotoImage(display_img)

                #t2_display = ImageTk.PhotoImage(t2_crop)
                t2_label = Label(t2_labelframe, image=t2_display)
                t2_label.image = t2_display
                t2_label.place(x=0,y=1)
                root.update()

                workbook = Workbook()
                sheet = workbook.active

                sheet["A2"] = "A"
                sheet["A3"] = "B"
                sheet["A4"] = "C"
                sheet["A5"] = "D"
                sheet["A6"] = "E"
                sheet["A7"] = "F"
                sheet["A8"] = "G"
                sheet["A9"] = "H"
                sheet["B1"] = "1"
                sheet["C1"] = "2"
                sheet["D1"] = "3"
                sheet["E1"] = "4"
                sheet["F1"] = "5"
                sheet["G1"] = "6"
                for i in range(0,48):
                    if(i<6):
                        pos = str(chr(65+i+1)) + "2"
                    if(i>=6 and i<12):
                        pos = str(chr(65+i-5)) + "3"
                    if(i>=12 and i<18):
                        pos = str(chr(65+i-11)) + "4"
                    if(i>=18 and i<24):
                        pos = str(chr(65+i-17)) + "5"
                    if(i>=24 and i<30):
                        pos = str(chr(65+i-23)) + "6"
                    if(i>=30 and i<36):
                        pos = str(chr(65+i-29)) + "7"
                    if(i>=36 and i<42):
                        pos = str(chr(65+i-35)) + "8"
                    if(i>=42):
                        pos = str(chr(65+i-41)) + "9"

                    if(id_list[i]=='N/A'):
                        sheet[pos] = 'N/A'
                    else:
                        sheet[pos] = t2_result[i]

                workbook.save(path3+"/T2.xlsx")

            if(receive_data=='C3'):
                t1_run=0
                t2_run=0
                t3_run=1
                print("Data received:", receive_data)
                t3wait_label.place_forget()
                t3_labelframe['bg'] = atk.DEFAULT_COLOR
                t3_labelframe['fg'] = 'lawn green'
                t_progressbar = atk.RadialProgressbar(t3_labelframe, fg='cyan')
                t_progressbar.place(x=47,y=70)
                t_progressbar.start()
                tprocess_label = Label(t3_labelframe, bg=atk.DEFAULT_COLOR, fg='white smoke', text='Đang phân tích...', font=("Courier",9,'bold'))
                tprocess_label.place(x=38,y=112)

                camera_capture(path1 + "/T3(1).jpg") 
                sleep(1)
                camera_capture(path1 + "/T3(2).jpg")

                send_data = 'C'
                ser.write(send_data.encode())
                print('Capture done!')
                #t3_result, t3_image = process_image(path1 + "/T3.jpg", start_point, end_point)
                t3_result1,_ = process_image(path1 + "/T3(1).jpg")
                t3_result2, t3_image = process_image(path1 + "/T3(2).jpg")

                t3_result = list(range(48))
                for i in range(0,48):
                    t3_result[i]=round((t3_result1[i]+t3_result2[i])/2,1)
                    if(t3_result[i] <= float(thr3l_set)):
                        cv2.drawContours(t3_image, sorted_contours1, i, (0,255,0), thickness = 2)
                    else:
                        cv2.drawContours(t3_image, sorted_contours1, i, (0,0,255), thickness = 2)

                output = path2 + "/T3.jpg"
                cv2.imwrite(output, t3_image)
                t3_analysis = Image.open(output)
                t3_crop = t3_analysis.crop((x1-13, y1-13, x2+13, y2+13))
                #t3_crop = t3_analysis.crop((280-7, 81-7, 498+7, 376+7))
                crop_width, crop_height = t3_crop.size
                scale_percent = 75
                width = int(crop_width * scale_percent / 100)
                height = int(crop_height * scale_percent / 100)
                display_img = t3_crop.resize((width,height))
                t3_display = ImageTk.PhotoImage(display_img)

                t3_label = Label(t3_labelframe, image=t3_display)
                t3_label.image = t3_display
                t3_label.place(x=0,y=1)
                wait = 0
                root.update()
                #pause_button.place_forget()
                stop_button.place_forget()
                temp_label.place_forget()
                autoprocess_label.place_forget()

                workbook = Workbook()
                sheet = workbook.active

                sheet["A2"] = "A"
                sheet["A3"] = "B"
                sheet["A4"] = "C"
                sheet["A5"] = "D"
                sheet["A6"] = "E"
                sheet["A7"] = "F"
                sheet["A8"] = "G"
                sheet["A9"] = "H"
                sheet["B1"] = "1"
                sheet["C1"] = "2"
                sheet["D1"] = "3"
                sheet["E1"] = "4"
                sheet["F1"] = "5"
                sheet["G1"] = "6"
                for i in range(0,48):
                    if(i<6):
                        pos = str(chr(65+i+1)) + "2"
                    if(i>=6 and i<12):
                        pos = str(chr(65+i-5)) + "3"
                    if(i>=12 and i<18):
                        pos = str(chr(65+i-11)) + "4"
                    if(i>=18 and i<24):
                        pos = str(chr(65+i-17)) + "5"
                    if(i>=24 and i<30):
                        pos = str(chr(65+i-23)) + "6"
                    if(i>=30 and i<36):
                        pos = str(chr(65+i-29)) + "7"
                    if(i>=36 and i<42):
                        pos = str(chr(65+i-35)) + "8"
                    if(i>=42):
                        pos = str(chr(65+i-41)) + "9"

                    if(id_list[i]=='N/A'):
                        sheet[pos] = 'N/A'
                    else:
                        sheet[pos] = t3_result[i]

                workbook.save(path3+"/T3.xlsx")

                t1_run=0
                t2_run=0
                t3_run=0

                if(server_on == 1):
                    workbook1 = load_workbook("/home/pi/Desktop/Spotcheck ID/" + excel_file, keep_vba = True)
                    sheet = workbook1.active
                else:
                    workbook1 = load_workbook("/home/pi/Spotcheck/template.xlsm", keep_vba = True)
                    sheet = workbook1.active

                sheet.protection.sheet = True
                sheet.protection.enable()

                if(server_on == 0):
                    sheet["C10"].protection = Protection(locked=False, hidden=False)
                #sheet["B7"].protection = Protection(locked=False, hidden=False)
                sheet["B8"].protection = Protection(locked=False, hidden=False)
                #sheet["B9"].protection = Protection(locked=False, hidden=False)

                font0 = Font(bold=False)
                font1 = Font(size='14', bold=True, color='00FF0000')
                font2 = Font(bold=True)
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                for i in range(12,60):
                    sheet["B"+str(i)].font = font0
                    sheet["D"+str(i)].font = font0

                img = Img('/home/pi/Spotcheck/logo.png')
                img.height = 39
                img.width = 215
                img.anchor = 'B2'
                sheet.add_image(img)

                sheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=6)
                sheet["B5"] = 'KẾT QUẢ CHẨN ĐOÁN COVID-19'
                sheet["B5"].font = font1
                sheet.cell(row=5,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                #global foldername
                sheet["B7"] = 'Tên tệp xét nghiệm: ' + importfilename
                sheet["B7"].font = font2
                sheet['B8'] = 'Người thực hiện: '
                sheet["B8"].font = font2
                #global covid19dir_old
                sheet['B9'] = 'Ngày thực hiện: ' + covid19dir_old[8:25]
                sheet["B9"].font = font2
                sheet['B60'] = 'Ghi chú:'
                sheet["B60"].font = font2
                sheet['B61'] = '+ N/A: Trống'
                sheet['B62'] = '+ E: Lỗi'
                sheet['C61'] = '+ R: Không xác định'
                sheet['C62'] = '+ N: Âm tính'
                sheet['E61'] = '+ P: Dương tính'

                sheet.merge_cells(start_row=64, start_column=4, end_row=64, end_column=6)
                sheet.merge_cells(start_row=65, start_column=4, end_row=65, end_column=6)
                sheet['B64'] = '​Kỹ thuật viên'
                sheet['B65'] = 'Ký tên'
                sheet['D64'] = '​Trưởng phòng xét nghiệm'
                sheet['D65'] = 'Ký tên'
                sheet["B64"].font = font2
                sheet["D64"].font = font2
                sheet["B64"].protection = Protection(locked=False, hidden=False)
                sheet["D64"].protection = Protection(locked=False, hidden=False)
                sheet.cell(row=64,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                sheet.cell(row=65,column=2).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                sheet.cell(row=64,column=4).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                sheet.cell(row=65,column=4).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)

                for r in range(11,60):
                    for c in range(2,7):
                        sheet.cell(row=r,column=c).alignment = Alignment(horizontal='center',vertical='center',wrapText=True)
                        sheet.cell(row=r,column=c).border = thin_border

                sheet.column_dimensions['B'].width = 26
                sheet.column_dimensions['C'].width = 12
                sheet.column_dimensions['D'].width = 12
                sheet.column_dimensions['E'].width = 12
                sheet.column_dimensions['F'].width = 12

                sheet.row_dimensions[11].height = 40

                sheet['B11'] = 'ID KHÁCH HÀNG'
                sheet["B11"].font = font2
                sheet["B11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
                sheet['C11'] = 'Vị trí mẫu'
                sheet["C11"].font = font2
                sheet["C11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
                sheet['D11'] = 'Kết quả Spotcheck'
                sheet["D11"].font = font2
                sheet["D11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
                sheet['E11'] = 'Kết quả Gel'
                sheet["E11"].font = font2
                sheet["E11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')
                sheet['F11'] = 'Kết luận'
                sheet["F11"].font = font2
                sheet["F11"].fill = PatternFill(start_color='00EFEFEF', end_color='00EFEFEF', fill_type='solid')

                for i in range (12,60):
                    if(i<20):
                        sheet['C'+str(i)] = str(chr(65+i-12)) + '1'
                    if(i>=20 and i<28):
                        sheet['C'+str(i)] = str(chr(65+i-20)) + '2'
                    if(i>=28 and i<36):
                        sheet['C'+str(i)] = str(chr(65+i-28)) + '3'
                    if(i>=36 and i<44):
                        sheet['C'+str(i)] = str(chr(65+i-36)) + '4'
                    if(i>=44 and i<52):
                        sheet['C'+str(i)] = str(chr(65+i-44)) + '5'
                    if(i>=52):
                        sheet['C'+str(i)] = str(chr(65+i-52)) + '6'

                c1=-6
                c2=-5
                c3=-4
                c4=-3
                c5=-2
                c6=-1
                for i in range(0,8):
                    c1=c1+6
                    sheet['B'+str(i+12)] = id_list[c1]
                    if(id_list[c1]=='N/A'):
                        sheet['D'+str(i+12)] = 'N/A'
                    else:
                        if(t1_result[c1]<=float(thr1_set)):
                            sheet['D'+str(i+12)] = 'E'
                            sheet['D'+str(i+12)].fill = PatternFill(start_color='00FFFF33', end_color='00FFFF33', fill_type='solid')
                        if(t1_result[c1]>float(thr1_set) and t2_result[c1]<=float(thr2_set) and t3_result[c1]<=float(thr3l_set)):
                            sheet['D'+str(i+12)] = 'N'
                            sheet['D'+str(i+12)].fill = PatternFill(start_color='0099FF00', end_color='0000FF00', fill_type='solid')
                        if(t1_result[c1]>float(thr1_set) and t2_result[c1]>float(thr2_set) and t3_result[c1]<=float(thr3l_set)):
                            sheet['D'+str(i+12)] = 'P'
                            sheet['D'+str(i+12)].fill = PatternFill(start_color='00FF4141', end_color='00FF4141', fill_type='solid')
                            sheet['D'+str(i+12)].font = font2
                            sheet['B'+str(i+12)].font = font2
                        if(t1_result[c1]>float(thr1_set) and t2_result[c1]>float(thr2_set) and t3_result[c1]>float(thr3l_set)):
                            sheet['D'+str(i+12)] = 'N'
                            sheet['D'+str(i+12)].fill = PatternFill(start_color='0099FF00', end_color='0000FF00', fill_type='solid')
                    
                    sheet['E'+str(i+12)].protection = Protection(locked=False, hidden=False)
                    sheet['F'+str(i+12)].protection = Protection(locked=False, hidden=False)

                    c2=c2+6
                    sheet['B'+str(i+20)] = id_list[c2]
                    if(id_list[c2]=='N/A'):
                        sheet['D'+str(i+20)] = 'N/A'
                    else:
                        if(t1_result[c2]<=float(thr1_set)):
                            sheet['D'+str(i+20)] = 'E'
                            sheet['D'+str(i+20)].fill = PatternFill(start_color='00FFFF33', end_color='00FFFF33', fill_type='solid')
                        if(t1_result[c2]>float(thr1_set) and t2_result[c2]<=float(thr2_set) and t3_result[c2]<=float(thr3l_set)):
                            sheet['D'+str(i+20)] = 'N'
                            sheet['D'+str(i+20)].fill = PatternFill(start_color='0099FF00', end_color='0000FF00', fill_type='solid')
                        if(t1_result[c2]>float(thr1_set) and t2_result[c2]>float(thr2_set) and t3_result[c2]<=float(thr3l_set)):
                            sheet['D'+str(i+20)] = 'P'
                            sheet['D'+str(i+20)].fill = PatternFill(start_color='00FF4141', end_color='00FF4141', fill_type='solid')
                            sheet['D'+str(i+20)].font = font2
                            sheet['B'+str(i+20)].font = font2
                        if(t1_result[c2]>float(thr1_set) and t2_result[c2]>float(thr2_set) and t3_result[c2]>float(thr3l_set)):
                            sheet['D'+str(i+20)] = 'N'
                            sheet['D'+str(i+20)].fill = PatternFill(start_color='0099FF00', end_color='0000FF00', fill_type='solid')
                       
                    sheet['E'+str(i+20)].protection = Protection(locked=False, hidden=False)
                    sheet['F'+str(i+20)].protection = Protection(locked=False, hidden=False)

                    c3=c3+6
                    sheet['B'+str(i+28)] = id_list[c3]
                    if(id_list[c3]=='N/A'):
                        sheet['D'+str(i+28)] = 'N/A'
                    else:
                        if(t1_result[c3]<=float(thr1_set)):
                            sheet['D'+str(i+28)] = 'E'
                            sheet['D'+str(i+28)].fill = PatternFill(start_color='00FFFF33', end_color='00FFFF33', fill_type='solid')
                        if(t1_result[c3]>float(thr1_set) and t2_result[c3]<=float(thr2_set) and t3_result[c3]<=float(thr3l_set)):
                            sheet['D'+str(i+28)] = 'N'
                            sheet['D'+str(i+28)].fill = PatternFill(start_color='0099FF00', end_color='0000FF00', fill_type='solid')
                        if(t1_result[c3]>float(thr1_set) and t2_result[c3]>float(thr2_set) and t3_result[c3]<=float(thr3l_set)):
                            sheet['D'+str(i+28)] = 'P'
                            sheet['D'+str(i+28)].fill = PatternFill(start_color='00FF4141', end_color='00FF4141', fill_type='solid')
                            sheet['D'+str(i+28)].font = font2
                            sheet['B'+str(i+28)].font = font2
                        if(t1_result[c3]>float(thr1_set) and t2_result[c3]>float(thr2_set) and t3_result[c3]>float(thr3l_set)):
                            sheet['D'+str(i+28)] = 'N'
                            sheet['D'+str(i+28)].fill = PatternFill(start_color='0099FF00', end_color='0000FF00', fill_type='solid')
                        
                    sheet['E'+str(i+28)].protection = Protection(locked=False, hidden=False)
                    sheet['F'+str(i+28)].protection = Protection(locked=False, hidden=False)

                    c4=c4+6
                    sheet['B'+str(i+36)] = id_list[c4]
                    if(id_list[c4]=='N/A'):
                        sheet['D'+str(i+36)] = 'N/A'
                    else:
                        if(t1_result[c4]<=float(thr1_set)):
                            sheet['D'+str(i+36)] = 'E'
                            sheet['D'+str(i+36)].fill = PatternFill(start_color='00FFFF33', end_color='00FFFF33', fill_type='solid')
                        if(t1_result[c4]>float(thr1_set) and t2_result[c4]<=float(thr2_set) and t3_result[c4]<=float(thr3l_set)):
                            sheet['D'+str(i+36)] = 'N'
                            sheet['D'+str(i+36)].fill = PatternFill(start_color='0099FF00', end_color='0000FF00', fill_type='solid')
                        if(t1_result[c4]>float(thr1_set) and t2_result[c4]>float(thr2_set) and t3_result[c4]<=float(thr3l_set)):
                            sheet['D'+str(i+36)] = 'P'
                            sheet['D'+str(i+36)].fill = PatternFill(start_color='00FF4141', end_color='00FF4141', fill_type='solid')
                            sheet['D'+str(i+36)].font = font2
                            sheet['B'+str(i+36)].font = font2
                        if(t1_result[c4]>float(thr1_set) and t2_result[c4]>float(thr2_set) and t3_result[c4]>float(thr3l_set)):
                            sheet['D'+str(i+36)] = 'N'
                            sheet['D'+str(i+36)].fill = PatternFill(start_color='0099FF00', end_color='0000FF00', fill_type='solid')
                        
                    sheet['E'+str(i+36)].protection = Protection(locked=False, hidden=False)
                    sheet['F'+str(i+36)].protection = Protection(locked=False, hidden=False)

                    c5=c5+6
                    sheet['B'+str(i+44)] = id_list[c5]
                    if(id_list[c5]=='N/A'):
                        sheet['D'+str(i+44)] = 'N/A'
                    else:
                        if(t1_result[c5]<=float(thr1_set)):
                            sheet['D'+str(i+44)] = 'E'
                            sheet['D'+str(i+44)].fill = PatternFill(start_color='00FFFF33', end_color='00FFFF33', fill_type='solid')
                        if(t1_result[c5]>float(thr1_set) and t2_result[c5]<=float(thr2_set) and t3_result[c5]<=float(thr3l_set)):
                            sheet['D'+str(i+44)] = 'N'
                            sheet['D'+str(i+44)].fill = PatternFill(start_color='0099FF00', end_color='0000FF00', fill_type='solid')
                        if(t1_result[c5]>float(thr1_set) and t2_result[c5]>float(thr2_set) and t3_result[c5]<=float(thr3l_set)):
                            sheet['D'+str(i+44)] = 'P'
                            sheet['D'+str(i+44)].fill = PatternFill(start_color='00FF4141', end_color='00FF4141', fill_type='solid')
                            sheet['D'+str(i+44)].font = font2
                            sheet['B'+str(i+44)].font = font2
                        if(t1_result[c5]>float(thr1_set) and t2_result[c5]>float(thr2_set) and t3_result[c5]>float(thr3l_set)):
                            sheet['D'+str(i+44)] = 'N'
                            sheet['D'+str(i+44)].fill = PatternFill(start_color='0099FF00', end_color='0000FF00', fill_type='solid')
                        
                    sheet['E'+str(i+44)].protection = Protection(locked=False, hidden=False)
                    sheet['F'+str(i+44)].protection = Protection(locked=False, hidden=False)

                    c6=c6+6
                    sheet['B'+str(i+52)] = id_list[c6]
                    if(id_list[c6]=='N/A'):
                        sheet['D'+str(i+52)] = 'N/A' 
                    else:
                        if(t1_result[c6]<=float(thr1_set)):
                            sheet['D'+str(i+52)] = 'E'
                            sheet['D'+str(i+52)].fill = PatternFill(start_color='00FFFF33', end_color='00FFFF33', fill_type='solid')
                        if(t1_result[c6]>float(thr1_set) and t2_result[c6]<=float(thr2_set) and t3_result[c6]<=float(thr3l_set)):
                            sheet['D'+str(i+52)] = 'N'
                            sheet['D'+str(i+52)].fill = PatternFill(start_color='0099FF00', end_color='0000FF00', fill_type='solid')
                        if(t1_result[c6]>float(thr1_set) and t2_result[c6]>float(thr2_set) and t3_result[c6]<=float(thr3l_set)):
                            sheet['D'+str(i+52)] = 'P'
                            sheet['D'+str(i+52)].fill = PatternFill(start_color='00FF4141', end_color='00FF4141', fill_type='solid')
                            sheet['D'+str(i+52)].font = font2
                            sheet['B'+str(i+52)].font = font2
                        if(t1_result[c6]>float(thr1_set) and t2_result[c6]>float(thr2_set) and t3_result[c6]>float(thr3l_set)):
                            sheet['D'+str(i+52)] = 'N'
                            sheet['D'+str(i+52)].fill = PatternFill(start_color='0099FF00', end_color='0000FF00', fill_type='solid')

                    sheet['E'+str(i+52)].protection = Protection(locked=False, hidden=False)
                    sheet['F'+str(i+52)].protection = Protection(locked=False, hidden=False)
                
                sheet.print_area = 'A1:G70'
                workbook1.save("/home/pi/Desktop/Ket Qua Phan Tich/" + importfilename + ".xlsm")

                if(os.path.exists("/home/pi/Desktop/Spotcheck ID/" + excel_file)):
                    try:
                        shutil.move("/home/pi/Desktop/Spotcheck ID/" + excel_file,"/home/pi/Desktop/Spotcheck ID/Spotcheck ID - Old")
                    except:
                        pass
                else:
                    pass

                if(server_on==1):
                    try:
                        ftp = FTP(ftp_ip, ftp_user, ftp_password)
                        ftp.cwd(ftp_folder + 'Processed_Data')
                        file = open("/home/pi/Desktop/Ket Qua Phan Tich/" + importfilename + ".xlsm",'rb')
                        ftp.storbinary('STOR ' + importfilename + ".xlsm", file)
                        ftp.quit()
                    except Exception as e :
                        error = messagebox.showwarning("Có lỗi xảy ra khi đồng bộ server !",str(e))
                        if(error=='ok'):
                            pass

                def thr():
                    th2 = Thread(target = viewresult_click)
                    th2.start()
                def viewresult_click():
                    viewresult_button.place_forget()
                    t1_labelframe.place_forget()
                    t2_labelframe.place_forget()
                    t3_labelframe.place_forget()
                    t_labelframe.place_forget()
                    analysis_label['text']="KẾT QUẢ PHÂN TÍCH"

                    annotate_labelframe = LabelFrame(analysis_labelframe, bg='white', width=380, height=305)
                    annotate_labelframe.place(x=360,y=76)
                    root.update_idletasks()

                    negative_label = Label(annotate_labelframe, bg='lawn green', width=4, height=2)
                    negative_label.place(x=75,y=32)
                    negativetext_label = Label(annotate_labelframe, bg='white', text='  (N)           ÂM TÍNH', height=2)
                    negativetext_label.place(x=145,y=32)
                    positive_label = Label(annotate_labelframe, bg='red', width=4, height=2)
                    positive_label.place(x=75,y=82)
                    positivetext_label = Label(annotate_labelframe, bg='white', text='  (P)           DƯƠNG TÍNH', height=2)
                    positivetext_label.place(x=145,y=82)
                    redue_label = Label(annotate_labelframe, bg='cyan', width=4, height=2)
                    redue_label.place(x=75,y=132)
                    reduetext_label = Label(annotate_labelframe, bg='white', text='  (R)           KHÔNG XÁC ĐỊNH', height=2)
                    reduetext_label.place(x=145,y=132)
                    none_label = Label(annotate_labelframe, bg='white smoke', width=4, height=2)
                    none_label.place(x=75,y=182)
                    nonetext_label = Label(annotate_labelframe, bg='white', text='(N/A)         TRỐNG', height=2)
                    nonetext_label.place(x=145,y=182)
                    error_label = Label(annotate_labelframe, bg='yellow', width=4, height=2)
                    error_label.place(x=75,y=232)
                    errortext_label = Label(annotate_labelframe, bg='white', text='  (E)           LỖI', height=2)
                    errortext_label.place(x=145,y=232)
                    root.update_idletasks()

                    result_labelframe = LabelFrame(analysis_labelframe, bg='ghost white', width=600,height = 307)
                    result_labelframe.place(x=104,y=120)
                    row_labelframe = LabelFrame(analysis_labelframe, bg='ghost white', width=600,height = 50)
                    row_labelframe.place(x=104,y=76)
                    column_labelframe = LabelFrame(analysis_labelframe, bg='ghost white', width=50,height = 307)
                    column_labelframe.place(x=62,y=120)
                    root.update_idletasks()

                    row_label = [0,0,0,0,0,0]
                    for i in range (0,6):
                        row_text = i+1
                        row_label[i] = Label(row_labelframe, text=row_text, bg='grey94', width=4, height=2)
                        row_label[i].grid(row=0,column=i,padx=2,pady=2)

                    column_label = [0,0,0,0,0,0,0,0]
                    for i in range (0,8):
                        if(i==0):
                            column_text = 'A'
                        if(i==1):
                            column_text = 'B'
                        if(i==2):
                            column_text = 'C'
                        if(i==3):
                            column_text = 'D'
                        if(i==4):
                            column_text = 'E'
                        if(i==5):
                            column_text = 'F'
                        if(i==6):
                            column_text = 'G'
                        if(i==7):
                            column_text = 'H'
                        column_label[i] = Label(column_labelframe, text=column_text, bg='grey94', width=4, height=2)
                        column_label[i].grid(row=i,column=0,padx=2,pady=2)

                    label = list(range(48))
                    def result_table(range_a,range_b, row_value):
                        j=-1
                        global pos_result
                        for i in range(range_a, range_b):
                            j+=1
                            if(id_list[i]=='N/A'):
                                label[i] = Label(result_labelframe, bg='white smoke', text='N/A', width=4, height=2)
                                label[i].grid(row=row_value,column=j,padx=2,pady=2)
                            else:
                                if(t1_result[i]<=float(thr1_set)):
                                    label[i] = Label(result_labelframe, bg='yellow', text='E', width=4, height=2)
                                    label[i].grid(row=row_value,column=j,padx=2,pady=2)
                                if(t1_result[i]>float(thr1_set) and t2_result[i]<=float(thr2_set) and t3_result[i]<=float(thr3l_set)):
                                    label[i] = Label(result_labelframe, bg='lawn green', text='N', width=4, height=2)
                                    label[i].grid(row=row_value,column=j,padx=2,pady=2)
                                if(t1_result[i]>float(thr1_set) and t2_result[i]>float(thr2_set) and t3_result[i]<=float(thr3l_set)):
                                    label[i] = Label(result_labelframe, bg='red', text='P', width=4, height=2)
                                    label[i].grid(row=row_value,column=j,padx=2,pady=2)
                                if(t1_result[i]>float(thr1_set) and t2_result[i]>float(thr2_set) and t3_result[i]>float(thr3l_set)):
                                    label[i] = Label(result_labelframe, bg='lawn green', text='N', width=4, height=2)
                                    label[i].grid(row=row_value,column=j,padx=2,pady=2)

                    result_table(0,6,0)
                    result_table(6,12,1)
                    result_table(12,18,2)
                    result_table(18,24,3)
                    result_table(24,30,4)
                    result_table(30,36,5)
                    result_table(36,42,6)
                    result_table(42,48,7)
                    
                    root.update_idletasks()

                    def detail_click():
                        if(detail_button['bg']=='lawn green'):
                            detail_button['bg']='grey94'
                            for i in range (0,48):
                                #if(pos_result[i]<=15):
                                if(id_list[i]=='N/A'):
                                    label[i]['text'] = 'N/A'

                                else:
                                    if(t1_result[i]<=float(thr1_set)):
                                    #if(t1_result[i]<=21.5):
                                        label[i]['text'] = 'E'
                                    if(t1_result[i]>float(thr1_set) and t2_result[i]<=float(thr2_set)):
                                    #if(t1_result[i]>21.5 and t2_result[i]<=21.5):
                                        label[i]['text'] = 'E'
                                    if(t1_result[i]>float(thr1_set) and t2_result[i]>float(thr2_set) and t3_result[i]<=float(thr3l_set)):
                                    #if(t1_result[i]>21.5 and t2_result[i]>21.5 and t3_result[i]<=21.5):
                                        label[i]['text'] = 'N'
                                    if(t1_result[i]>float(thr1_set) and t2_result[i]>float(thr2_set) and t3_result[i]>float(thr3l_set) and t3_result[i]<=float(thr3h_set)):
                                    #if(t1_result[i]>21.5 and t2_result[i]>21.5 and t3_result[i]>21.5 and t3_result[i]<=21.5):
                                        label[i]['text'] = 'R'
                                    if(t1_result[i]>float(thr1_set) and t2_result[i]>float(thr2_set) and t3_result[i]>float(thr3h_set)):
                                    #if(t1_result[i]>21.5 and t2_result[i]>21.5 and t3_result[i]>22.5):
                                        label[i]['text'] = 'P'
                        else:
                            detail_button['bg']='lawn green'
                            for i in range (0,48):
                                #if(pos_result[i]<=15):
                                if(id_list[i]=='N/A'):
                                    label[i]['text'] = str('%.1f'%t3_result[i])

                                else:
                                    if(t1_result[i]<=float(thr1_set)):
                                    #if(t1_result[i]<=21.5):
                                        label[i]['text'] = str('%.1f'%t3_result[i])
                                    if(t1_result[i]>float(thr1_set) and t2_result[i]<=float(thr2_set)):
                                    #if(t1_result[i]>21.5 and t2_result[i]<=21.5):
                                        label[i]['text'] = str('%.1f'%t3_result[i])
                                    if(t1_result[i]>float(thr1_set) and t2_result[i]>float(thr2_set) and t3_result[i]<=float(thr3l_set)):
                                    #if(t1_result[i]>21.5 and t2_result[i]>21.5 and t3_result[i]<=21.5):
                                        label[i]['text'] = str('%.1f'%t3_result[i])
                                    if(t1_result[i]>float(thr1_set) and t2_result[i]>float(thr2_set) and t3_result[i]>(float(thr3l_set)+plus_value1) and t3_result[i]<=(float(thr3l_set)+plus_value2)):
                                    #if(t1_result[i]>21.5 and t2_result[i]>21.5 and t3_result[i]>21.5 and t3_result[i]<=21.5):
                                        label[i]['text'] = str('%.1f'%t3_result[i])
                                    if(t1_result[i]>float(thr1_set) and t2_result[i]>float(thr2_set) and t3_result[i]>float(thr3l_set) and t3_result[i]<=(float(thr3l_set)+plus_value1)):
                                    #if(t1_result[i]>21.5 and t2_result[i]>21.5 and t3_result[i]>22.5):
                                        label[i]['text'] = str('%.1f'%t3_result[i])
                                    if(t1_result[i]>float(thr1_set) and t2_result[i]>float(thr2_set) and t3_result[i]>(float(thr3l_set)+plus_value2)):
                                        label[i]['text'] = str('%.1f'%t3_result[i])

                            root.update_idletasks()
                            subprocess.call(["scrot",path0+"/gia-tri.jpg"])

                    def finish_click():
                        msgbox = messagebox.askquestion('Ket thuc chuong trinh','Bạn có muốn quay lại ?', icon = 'question')
                        if(msgbox=='yes'):
                            for i in range (0,48):
                                label[i]['text'] = str('%.1f'%t3_result[i])                         
                            root.update_idletasks()
                            sleep(1)
                            subprocess.call(["scrot",path3+"/gia-tri.jpg"])
                            sleep(1)
                            global foldername
                            global covid19clicked
                            foldername = ""
                            covid19clicked = 1
                            analysis_labelframe.place_forget()
                            global wait
                            wait=0
                            mainscreen()

                    # detail_button = Button(analysis_labelframe, activebackground="white", bg="grey94", text="Chi tiết", height=3, width=10, borderwidth=0, command=detail_click)
                    # detail_button.place(x=360,y=396)
                    finish_button = Button(analysis_labelframe, bg="dark orange", text="Hoàn thành", height=3, width=15, borderwidth=0, command=finish_click)
                    finish_button.place(x=480,y=396)

                    root.update_idletasks()

                    subprocess.call(["scrot",path0+"/ket-qua.jpg"])
#                     screenshot_img = Image.open(path3+"/result.jpg")
#                     screenshot_crop = screenshot_img.crop((60,74,352,475))
#                     screenshot_crop = screenshot_crop.save(path3+"/result.jpg")

                viewresult_button = Button(analysis_labelframe, bg="dodger blue", text="Kết quả", height=3, width=15, borderwidth=0, command=thr)
                viewresult_button.place(x=327,y=394)
########################################################## SAMPLES ANALYSIS - END ##################################################################

############################################################## WARNING - START #####################################################################
# def warning(channel):
#     global warning_value
#     if(warning_value==1):
#         warning_label = Label(mainscreen_labelframe, bg='white', fg='white', text='Hệ thống đang tản nhiệt, không đặt mẫu vào lúc này !', font=("Courier", 13, 'bold'))
#         warning_label.place(x=220,y=450)
#         warning_value = 0
#         print("Warning:", warning_value)
#     else:
#         warning_label = Label(mainscreen_labelframe, bg='red',text='Hệ thống đang tản nhiệt, không đặt mẫu vào lúc này !', font=("Courier", 13, 'bold'))
#         warning_label.place(x=220,y=450)
#         warning_value = 1
#         print("Warning:", warning_value)

# GPIO.setmode(GPIO.BCM)
# GPIO.setup(16, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
# GPIO.add_event_detect(16,GPIO.FALLING,callback=warning)
############################################################### WARNING - END ######################################################################

############################################################### LOOP - START #######################################################################
# ser.flushInput()
# ser.flushOutput()
# send_data = 'o'
# ser.write(send_data.encode())
while True:
    if(start_trial==1):
        trial()
    else:
        mainscreen()
    root.mainloop()
################################################################ LOOP - END ########################################################################
