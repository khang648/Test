#!/usr/bin/python3
########################################################### IMPORT MODULES - START #################################################################
from tkinter import *
from tkinter import messagebox
import time
from time import sleep, gmtime, strftime
from picamera import PiCamera
import cv2
import numpy as np
from tkinter import filedialog
from PIL import ImageTk, Image
import serial
from functools import partial
import math
from fractions import Fraction
from threading import *
import os
from tkinter import ttk
import awesometkinter as atk
import tkinter.font as font
import openpyxl
import subprocess
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image as Img
from datetime import *
############################################################ IMPORT MODULES - END ##################################################################

########################################################## GLOBAL VARIABLE - START #################################################################
name = "/"
entry_num = 0
wait = 0
pos_result = list(range(48))
path0 = "/"
path1 = "/"
path2 = "/"
path3 = "/"
path4 = "/"
path5 = "/"
foldername = ""
importfilename = ""
id_list = list(range(48))
samples = 0
div = list(range(48))
start_point = (0,0)
end_point = (0,0)
thr_set= 15
tmp = 0
password = '123456789'
value_max = 30
value_min = 10


fr = open("/home/pi/Spotcheck/check.txt","r")
code = (fr.readline()).strip()
fr1 = open("/home/pi/Spotcheck/coordinates1.txt","r")
x1 = int(fr1.readline())
y1 = int(fr1.readline())
x2 = int(fr1.readline())
y2 = int(fr1.readline())
fr2 = open("/var/tmp/.admin.txt","r")
start_trial = int(fr2.readline())
print("start_trial: ", start_trial)
fr3 = open("/home/pi/Spotcheck/mmvalue.txt")
value_min = float(fr3.readline())
value_max = float(fr3.readline())
fr4 = open("/home/pi/Spotcheck/threshold.txt","r")
raw = float(fr4.readline())
threshold = float(fr4.readline())
ratio1 = float(fr4.readline())
ratio2 = float(fr4.readline())
########################################################### GLOBAL VARIABLE - END ##################################################################

############################################################### TRIAL - START ######################################################################
def trial():
    old_day = int(fr2.readline())
    old_month = int(fr2.readline())
    old_year = int(fr2.readline())
    limit = int(fr2.readline())
    print("Ngày bắt đầu:", old_day, old_month, old_year)

    today = datetime.now()
    new_day = today.day
    new_month = today.month
    new_year = today.year
    print("Ngày hiện tại:", new_day, new_month, new_year)

    nam = new_year - old_year
    if(new_month < old_month):
        thang = new_month + 12 - old_month
        nam = nam - 1
    else:
        thang = new_month - old_month
    if(new_day < old_day):
        ngay = new_day + 30 - old_day
        thang = thang - 1
    else:
        ngay = new_day - old_day
    songay = ngay + thang*30 + nam*365
    print("Thời gian dùng thử còn lại:", limit-songay)
    if(songay >= limit):
        trial_labelframe = LabelFrame(root, bg='white', width=800, height=600)
        trial_labelframe.place(x=0,y=0)

        logo_img = Image.open('/home/pi/Spotcheck/logo.png')
        logo_width, logo_height = logo_img.size
        scale_percent = 50
        width = int(logo_width * scale_percent / 100)
        height = int(logo_height * scale_percent / 100)
        display_img = logo_img.resize((width,height))
        image_select = ImageTk.PhotoImage(display_img)
        logo_label = Label(trial_labelframe, bg='white',image=image_select)
        logo_label.image = image_select
        logo_label.place(x=5,y=5)

        def active_click(event = None):
            code = activecode_entry.get()
            if(code==""):
                msg = messagebox.showwarning(" ","Bạn chưa nhập mã kích hoạt !")
            else:
                if(code!=password):
                    msg = messagebox.showerror(" ","Mã kích hoạt không đúng !")
                if(code==password):
                    msg = messagebox.showinfo(" ","Kích hoạt thành công !")
                    f1=open("/var/tmp/admin.txt",'w')
                    f1.writelines("0")
                    mainscreen()

        trial_label = Label(trial_labelframe, bg='white',fg="red", text="Thời gian dùng thử đã hết\nVui lòng nhập mã kích hoạt để tiếp tục sử dụng !", font=("Courier",18,"bold"))
        trial_label.place(x=62,y=85)
        contact_label = Label(trial_labelframe, bg='white', text="Liên hệ nhà cung cấp để nhận mã kích hoạt:", font=("Courier",12,"bold"))
        contact_label.place(x=73,y=435)
        mail_label = Label(trial_labelframe, bg='white', fg='blue',text="cskh@phusabiochem.com", font=("Courier",12,"bold"))
        mail_label.place(x=503,y=435)
        activecode_entry = Entry(trial_labelframe, width=27, bg='white', font=("Courier",14,"bold"))
        activecode_entry.place(x=246,y=215)
        activecode_entry.bind("<Return>", active_click)
        code_label = Label(trial_labelframe, bg='white', text="Mã kích hoạt:", font=("Courier",14,"bold"))
        code_label.place(x=244,y=189)

        key_img = Image.open('/home/pi/Spotcheck/key.png')
        logo_width, logo_height = key_img.size
        scale_percent = 5
        width = int(logo_width * scale_percent / 100)
        height = int(logo_height * scale_percent / 100)
        display_img = key_img.resize((width,height))
        image_select = ImageTk.PhotoImage(display_img)
        logo_label = Label(trial_labelframe, bg='white',image=image_select)
        logo_label.image = image_select
        logo_label.place(x=726,y=85)

        active_button = Button(trial_labelframe, bg="lavender", font=("Courier",11,'bold'), text="Xác nhận", height=3, width=10, borderwidth=0, command=active_click)
        active_button.place(x=340,y=260)
    else:
        mainscreen()
################################################################ TRIAL - END #######################################################################

########################################################## MAIN WINDOW INIT - START ################################################################
root = Tk()
root.title(" ")
root.geometry('1024x600')
root.configure(background = "white")
root.attributes('-fullscreen', True)
root.resizable(False,False)
def disable_event():
    pass
root.protocol("WM_DELETE_WINDOW", disable_event)
s = ttk.Style()
s.theme_use('clam')
########################################################### MAIN WINDOW INIT - END #################################################################

############################################################ CAMERA INIT - START ###################################################################
def camera_capture(output):
    camera = PiCamera(framerate=Fraction(1,6), sensor_mode=3)
    camera.rotation = 180
    camera.iso = 200
    sleep(2)
    camera.shutter_speed = 6000000
    camera.exposure_mode = 'off'
    camera.capture(output)
    camera.close()
############################################################# CAMERA INIT - END ####################################################################

############################################################ SERIAL INIT - START ###################################################################
ser = serial.Serial(
    port = '/dev/serial0',
    baudrate = 115200,
    parity = serial.PARITY_NONE,
    stopbits = serial.STOPBITS_ONE,
    bytesize = serial.EIGHTBITS,
    timeout = 1
)

############################################################# SERIAL INIT - END ####################################################################

######################################################### SORTING CONTOURS - START #################################################################
def sorting_y(contour):
    rect_y = cv2.boundingRect(contour)
    return rect_y[1]
def sorting_x(contour):
    rect_x = cv2.boundingRect(contour)
    return rect_x[0]
def sorting_xy(contour):
    rect_xy = cv2.boundingRect(contour)
    return math.sqrt(math.pow(rect_xy[0],2) + math.pow(rect_xy[1],2))
########################################################## SORTING CONTOURS - END ##################################################################

########################################################## IMAGE ANALYSIS - START ##################################################################
def process_image(image_name, start_point=(x1,y1), end_point=(x2,y2)):
    image = cv2.imread(image_name)
    blur_img = cv2.GaussianBlur(image.copy(), (35,35), 0)
    gray_img = cv2.cvtColor(blur_img, cv2.COLOR_BGR2GRAY)
    thresh, binary_img = cv2.threshold(gray_img.copy(), 30, maxval=255, type=cv2.THRESH_BINARY)
    contours, hierarchy = cv2.findContours(binary_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    print("Number of contours: " + str(len(contours)))

    contours.sort(key=lambda data:sorting_xy(data))

    contour_img = np.zeros_like(gray_img)
    contour_img = cv2.rectangle(contour_img, start_point, end_point, (255,255,255), -1)
    rect_w = end_point[0] - start_point[0]
    rect_h = end_point[1] - start_point[1]
    cell_w = round(rect_w/6)
    cell_h = round(rect_h/8)
    for i in range(1,6):
        contour_img = cv2.line(contour_img, (start_point[0]+i*cell_w,start_point[1]), (start_point[0]+i*cell_w,end_point[1]),(0,0,0), 4)
    for i in range(1,8):
        contour_img = cv2.line(contour_img, (start_point[0],start_point[1]+i*cell_h), (end_point[0],start_point[1]+i*cell_h),(0,0,0), 4)

    thresh1 , binary1_img = cv2.threshold(contour_img, 250, maxval=255, type=cv2.THRESH_BINARY)
    contours1, hierarchy1 = cv2.findContours(binary1_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)

    contours1.sort(key=lambda data:sorting_y(data))
    contours1_h1 = contours1[0:6]
    contours1_h2 = contours1[6:12]
    contours1_h3 = contours1[12:18]
    contours1_h4 = contours1[18:24]
    contours1_h5 = contours1[24:30]
    contours1_h6 = contours1[30:36]
    contours1_h7 = contours1[36:42]
    contours1_h8 = contours1[42:48]
    contours1_h1.sort(key=lambda data:sorting_x(data))
    contours1_h2.sort(key=lambda data:sorting_x(data))
    contours1_h3.sort(key=lambda data:sorting_x(data))
    contours1_h4.sort(key=lambda data:sorting_x(data))
    contours1_h5.sort(key=lambda data:sorting_x(data))
    contours1_h6.sort(key=lambda data:sorting_x(data))
    contours1_h7.sort(key=lambda data:sorting_x(data))
    contours1_h8.sort(key=lambda data:sorting_x(data))

    sorted_contours1 = contours1_h1 + contours1_h2 + contours1_h3 + contours1_h4 + contours1_h5 + contours1_h6 + contours1_h7 + contours1_h8

    list_intensities = []
    sum_intensities = []
    result_list = list(range(48))
    area = list(range(48))

    blur1_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
    tmp_list = list(range(48))
    list_bgrvalue = []
    list_index = list(range(48))
    for i in range(len(sorted_contours1)):
        list_index[i] = []
        cimg = np.zeros_like(gray_img)
        cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
        pts = np.where(cimg == 255)
        list_bgrvalue.append(blur1_img[pts[0], pts[1]])
        for j in range(len(list_bgrvalue[i])):
             list_index[i].append(round((list_bgrvalue[i][j][0]+list_bgrvalue[i][j][1]+list_bgrvalue[i][j][2]/3)))
        list_index[i].sort()
        list_intensities.append(sum(list_index[i][len(list_index[i])-250:]))
        area[i]= cv2.contourArea(sorted_contours1[i])
        tmp_list[i] = list_intensities[i]/1000
        result_list[i] = round(tmp_list[i],1)

    for i in range(len(sorted_contours1)):
        if(result_list[i]>99):
            result_list[i]=99

    for i in range(len(sorted_contours1)):
        if ((i!=0) and ((i+1)%6==0)):
            print('%.1f'%(result_list[i]))
        else:
            print('%.1f'%(result_list[i]), end = ' | ')

    blurori_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
    global thr_set, id_list
    for i in range(len(sorted_contours1)):
        if(id_list[i]=='N/A'):
            cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,0), thickness = -1)
        else:
            if(result_list[i]<=10):
                cv2.drawContours(blurori_img, sorted_contours1, i, (255,255,0), thickness = 2)
            else:
                if(result_list[i] <= float(thr_set)):
                    cv2.drawContours(blurori_img, sorted_contours1, i, (255,255,0), thickness = 2)
                else:
                    cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,255), thickness = 2)

    return (result_list, blurori_img)
########################################################### IMAGE ANALYSIS - END ###################################################################

############################################################ MAIN SCREEN - START ###################################################################
def mainscreen():
    buttonFont = font.Font(family='Helvetica', size=10, weight='bold')
    global mainscreen_labelframe
    mainscreen_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    mainscreen_labelframe.place(x=0,y=0)

    logo_img = Image.open('/home/pi/Spotcheck/logo.png')
    logo_width, logo_height = logo_img.size
    scale_percent = 50
    width = int(logo_width * scale_percent / 100)
    height = int(logo_height * scale_percent / 100)
    display_img = logo_img.resize((width,height))
    image_select = ImageTk.PhotoImage(display_img)
    logo_label = Label(mainscreen_labelframe, bg='white',image=image_select)
    logo_label.image = image_select
    logo_label.place(x=5,y=5)

    sc_label = Label(mainscreen_labelframe, font=("Courier",45,'bold'), fg='dodger blue', bg='white',text='SPOTCHECK')
    sc_label.place(x=227, y=65)
    lightcheck_label = Label(mainscreen_labelframe, font=("Courier",30,'bold'), fg='grey40', bg='white',text='Kiểm tra hệ thống')
    lightcheck_label.place(x=188, y=135)

#     process_label = Label(mainscreen_labelframe, text='Đang xử lý...', bg='white', font=("Courier",13,'bold'))
#     process_label.place(x=240,y=350)
#     scanposition_progressbar = ttk.Progressbar(mainscreen_labelframe, orient = HORIZONTAL, style="green.Horizontal.TProgressbar", length = 200, mode = 'determinate')
#     scanposition_progressbar.place(x=204,y=310)
#     scanposition_progressbar['value'] = 5

    def start_click():
        send_data = 'P'
        ser.write(send_data.encode())

        process_label = Label(mainscreen_labelframe, text='Đang xử lý...', bg='white', font=("Courier",13,'bold'))
        process_label.place(x=330,y=350)
        root.update_idletasks()
        scanposition_progressbar = ttk.Progressbar(mainscreen_labelframe, orient = HORIZONTAL, style="green.Horizontal.TProgressbar", length = 200, mode = 'determinate')
        scanposition_progressbar.place(x=290,y=310)
        scanposition_progressbar['value'] = 5
        root.update_idletasks()
        start_button.place_forget()

        if(ser.in_waiting>0):
            receive_data = ser.readline().decode('utf-8').rstrip()
            print("Data received:", receive_data)
            scanposition_progressbar['value'] = 20
            root.update_idletasks()
            if(receive_data=='C'):
                global wait
                wait = 1

        while(wait!=1):
            root.update_idletasks()
            if(ser.in_waiting>0):
                receive_data = ser.readline().decode('utf-8').rstrip()
                print("Data received:", receive_data)
                if(receive_data=='C'):
                    wait = 1
                    break
        while(wait==1):
            try:
                camera_capture('/home/pi/Spotcheck/Kiem tra do sang/do-sang.jpg')
            except Exception as e :
                error = messagebox.askquestion("Lỗi: "+ str(e), "Bạn có muốn thoát chương trình ?", icon = "error")
                if(error=='yes'):
                    root.destroy()

            global test_list
            try:
                test_list,test_img = process_image('/home/pi/Spotcheck/Kiem tra do sang/do-sang.jpg')
            except Exception as e :
                error = messagebox.askquestion("Lỗi: "+ str(e), "Bạn có muốn thoát chương trình ?", icon = "error")
                if(error=='yes'):
                    root.destroy()
            cv2.imwrite('/home/pi/Spotcheck/Kiem tra do sang/xu-ly.jpg',test_img)
            workbook = Workbook()
            sheet = workbook.active

            sheet["A2"] = "A"
            sheet["A3"] = "B"
            sheet["A4"] = "C"
            sheet["A5"] = "D"
            sheet["A6"] = "E"
            sheet["A7"] = "F"
            sheet["A8"] = "G"
            sheet["A9"] = "H"
            sheet["B1"] = "1"
            sheet["C1"] = "2"
            sheet["D1"] = "3"
            sheet["E1"] = "4"
            sheet["F1"] = "5"
            sheet["G1"] = "6"
            for i in range(0,48):
                if(i<6):
                    pos = str(chr(65+i+1)) + "2"
                if(i>=6 and i<12):
                    pos = str(chr(65+i-5)) + "3"
                if(i>=12 and i<18):
                    pos = str(chr(65+i-11)) + "4"
                if(i>=18 and i<24):
                    pos = str(chr(65+i-17)) + "5"
                if(i>=24 and i<30):
                    pos = str(chr(65+i-23)) + "6"
                if(i>=30 and i<36):
                    pos = str(chr(65+i-29)) + "7"
                if(i>=36 and i<42):
                    pos = str(chr(65+i-35)) + "8"
                if(i>=42):
                    pos = str(chr(65+i-41)) + "9"

                sheet[pos] = test_list[i]

            workbook.save('/home/pi/Spotcheck/Kiem tra do sang/do-sang.xlsx')

            scanposition_progressbar['value'] = 50
            root.update_idletasks()
            sleep(1)
            scanposition_progressbar['value'] = 70
            root.update_idletasks()
            sleep(1)
            scanposition_progressbar['value'] = 99
            root.update_idletasks()

            process_label.place_forget()
            scanposition_progressbar.place_forget()
            #start_button.place(x=315,y=250)
            wait = 0
            break

        if(test_list[20]>value_max or test_list[21]>value_max or test_list[26]>value_max or test_list[27]>value_max or
           test_list[20]<value_min or test_list[21]<value_min or test_list[26]<value_min or test_list[27]<value_min):
            msgbox = messagebox.showerror(" ","Hệ thống lỗi, vui lòng liên hệ với nhà cung cấp !")
            fw = open("/home/pi/Spotcheck/check.txt","w")
            fw.truncate(0)
            fw.writelines("1111\n")
            if(msgbox=='ok'):
                root.destroy()
        else:
            global raw, threshold, ratio1, ratio2
            average_value = round(sum(test_list)/len(test_list),1)
            print("average_value:", average_value)
            #thr3l_value =  round((threshold + (average_value - 0)/(ratio1/ratio2)),1)
            tmp_value = int((average_value - raw)/ratio1)
            thr2_value =  round((tmp_value*ratio2)+threshold, 1)
            #thr3h_value = thr3l_value + 0.2
            thr3l_value = round(thr2_value/1.05, 1)
            thr1_value = thr3l_value - 1

            fw0 = open("/home/pi/Spotcheck/ct.txt","w")
            fw0.truncate(0)
            fw0.writelines(str(thr1_value)+"\n")
            fw0.writelines(str(thr2_value)+"\n")
            fw0.writelines(str(thr3l_value)+"\n")
            #fw0.writelines(str(thr3h_value)+"\n")

            fw = open("/home/pi/Spotcheck/check.txt","w")
            fw.truncate(0)
            fw.writelines("1234\n")
            msgbox = messagebox.showinfo(" ","Thiết bị đã sẵn sàng sử dụng !")
            if(msgbox == 'ok'):
                root.destroy()

    start_button = Button(mainscreen_labelframe, bg="grey98", text="Bắt đầu", font=('Courier',12,'bold'), borderwidth=0, height=3, width=12, command=start_click)
    #start_button.place(x=350,y=250)

    if(tmp==0):
        send_data = 'H'
        ser.write(send_data.encode())
        #msgbox = messagebox.showwarning("","Hệ thống sẽ tiến hành gia nhiệt tự động\nVui lòng chờ trong vài phút !")
        sc_label = Label(mainscreen_labelframe, font=("Courier",15,'bold'), fg ='red', bg='white',text='Hệ thống sẽ tiến hành gia nhiệt tự động\nVui lòng chờ trong vài phút...')
        sc_label.place(x=158, y=260)
    else: 
        try:
            sc_label.place_forget()
        except:
            pass
        start_click()

############################################################### LOOP - START #######################################################################
def readSerial():
    if(ser.in_waiting>0):
        receive_data = ser.readline().rstrip()
        print("Data received:", receive_data)
        if(receive_data==b'F\xff'):
            global tmp
            tmp=tmp+1
            if(tmp<2):
                mainscreen()
    root.after(100, readSerial)

if(start_trial==1):
    trial()
else:
    mainscreen()

root.after(100, readSerial)
root.mainloop()
################################################################ LOOP - END ########################################################################